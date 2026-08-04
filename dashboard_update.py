#!/usr/bin/env python3
"""
IQAM Dashboard – tägliche Aktualisierung
Läuft in GitHub Actions, authentifiziert via Microsoft Graph (Refresh Token).
Liest INVENTARBLATT + INVENTARLISTE aus den letzten Mails,
generiert dashboard_data.json + IQAM_Dashboard.html,
committed die Dateien in den docs/ Ordner.
"""

import json
import os
import sys
import tempfile
import base64
import re
import copy
from datetime import datetime, date
from urllib.request import Request, urlopen
from urllib.error import HTTPError
from urllib.parse import urlencode
import traceback

try:
    import requests as _requests
    _HAS_REQUESTS = True
except ImportError:
    _HAS_REQUESTS = False

# ─── Konfiguration ────────────────────────────────────────────────────────────
SENDER_EMAIL = "rbi-fondsreporting@rbinternational.com"
FUNDS = [
    {"id": "3411", "isin": "AT0000A1QA38", "color": "#F97316",
     "name": "Standortfonds AT"},
    {"id": "3431", "isin": "AT0000A1Z882", "color": "#3B7DD8",
     "name": "Standortfonds DE"},
    {"id": "3581", "isin": "AT0000A3EAW0", "color": "#16A34A",
     "name": "Dividends and Interest"},
]

# ─── Microsoft Graph: Token holen ─────────────────────────────────────────────
def get_access_token():
    client_id     = os.environ["MS_CLIENT_ID"]
    tenant_id     = os.environ["MS_TENANT_ID"]
    refresh_token = os.environ["MS_REFRESH_TOKEN"]

    data = urlencode({
        "grant_type":    "refresh_token",
        "client_id":     client_id,
        "refresh_token": refresh_token,
        "scope":         "https://graph.microsoft.com/Mail.Read offline_access",
    }).encode()

    req = Request(
        f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token",
        data=data,
        method="POST",
    )
    with urlopen(req) as resp:
        result = json.loads(resp.read())

    if "access_token" not in result:
        print(f"❌ Token-Fehler: {result.get('error_description', result)}")
        sys.exit(1)

    print("✅ Access token erhalten")
    return result["access_token"]


# ─── Graph API Helper ─────────────────────────────────────────────────────────
def graph_get(access_token, path):
    url = f"https://graph.microsoft.com/v1.0{path}"
    headers = {"Authorization": f"Bearer {access_token}", "Accept": "application/json"}
    if _HAS_REQUESTS:
        resp = _requests.get(url, headers=headers)
        resp.raise_for_status()
        return resp.json()
    else:
        from urllib.parse import quote
        safe_url = quote(url, safe="/:?&=.$,@'_-+%")
        req = Request(safe_url, headers=headers)
        with urlopen(req) as r:
            return json.loads(r.read())


def graph_get_bytes(access_token, path):
    url = f"https://graph.microsoft.com/v1.0{path}"
    headers = {"Authorization": f"Bearer {access_token}"}
    if _HAS_REQUESTS:
        resp = _requests.get(url, headers=headers)
        resp.raise_for_status()
        return resp.content
    else:
        from urllib.parse import quote
        safe_url = quote(url, safe="/:?&=.$,@'_-+%")
        req = Request(safe_url, headers=headers)
        with urlopen(req) as r:
            return r.read()


# ─── Mail + Attachment suchen ─────────────────────────────────────────────────
def graph_get_url(access_token, url):
    """Holt eine vollständige URL (für Pagination mit @odata.nextLink)."""
    headers = {"Authorization": f"Bearer {access_token}", "Accept": "application/json"}
    if _HAS_REQUESTS:
        resp = _requests.get(url, headers=headers)
        resp.raise_for_status()
        return resp.json()
    else:
        from urllib.parse import quote
        safe_url = quote(url, safe="/:?&=.$,@'_-+%")
        req = Request(safe_url, headers=headers)
        with urlopen(req) as r:
            return json.loads(r.read())


def find_all_historical_emails(access_token):
    """Holt ALLE historischen Mails vom Fonds-Sender (mit Pagination, manueller Filterung)."""
    url = (
        "https://graph.microsoft.com/v1.0/me/messages"
        "?$top=100"
        "&$orderby=receivedDateTime desc"
        "&$select=id,subject,receivedDateTime,hasAttachments,from"
    )
    all_messages = []
    page = 0
    while url:
        page += 1
        print(f"  📧 Seite {page}: {len(all_messages)} Sender-Mails bisher…")
        try:
            data = graph_get_url(access_token, url)
        except Exception as e:
            print(f"  ⚠️  Fehler bei Seite {page}: {e}")
            break
        msgs = data.get("value", [])
        # Manuell nach Sender filtern
        for msg in msgs:
            sender = msg.get("from", {}).get("emailAddress", {}).get("address", "").lower()
            if sender == SENDER_EMAIL.lower():
                all_messages.append(msg)
        url = data.get("@odata.nextLink")
    print(f"  ✅ {len(all_messages)} Mails vom Fonds-Sender gefunden")
    return all_messages


def backfill_nav_history_from_emails(access_token, existing_nav_history):
    """Liest alle historischen INVENTARBLATT-Mails und befüllt nav_history mit täglichen Preisen."""
    messages = find_all_historical_emails(access_token)
    nav_history = {k: list(v) for k, v in existing_nav_history.items()}  # shallow copy

    # Gruppiere nach Fund + Eingangsdatum
    mail_map = {}  # {(fid, recv_date): msg_id}
    for msg in messages:
        subj_up = msg.get("subject", "").upper()
        if "INVENTARBLATT" not in subj_up:
            continue
        recv_date = msg.get("receivedDateTime", "")[:10]
        for fund in FUNDS:
            fid = fund["id"]
            if fid not in msg.get("subject", ""):
                continue
            key = (fid, recv_date)
            if key not in mail_map:
                mail_map[key] = msg["id"]

    total = len(mail_map)
    print(f"\n📋 {total} historische INVENTARBLATT-Mails gefunden")

    # Alle Emails chronologisch verarbeiten — keine Skip-Logik.
    # Die finale Deduplizierung (letzter Eintrag pro Datum gewinnt) sorgt dafür,
    # dass spätere Emails (mit den offiziellen Closing-Werten) frühere überschreiben.
    # Beispiel: 01.07-Email enthält den offiziellen 30.06-Closing (164.16M) und
    # überschreibt den vorläufigen Wert der 30.06-Email (163.86M).
    done = 0
    new_entries = []  # alle neu geparseten Einträge (inkl. Überschreibungen)

    for (fid, recv_date), msg_id in sorted(mail_map.items()):
        done += 1
        print(f"  [{done}/{total}] Lade {fid} {recv_date}…")
        try:
            xlsx_bytes, filename = download_attachment(access_token, msg_id, ".xlsx")
            if not xlsx_bytes:
                print(f"    ⚠️  Kein Anhang")
                continue

            blatt_data = parse_excel(xlsx_bytes, fid)
            price       = blatt_data.get("nav_per_share")
            nav         = blatt_data.get("nav")
            shares      = blatt_data.get("shares")
            perf_ytd    = blatt_data.get("perf_ytd")
            asset_date  = blatt_data.get("asset_date")   # Datum aus "Asset (by DD.MM.YYYY)"-Zeile
            report_date = blatt_data.get("report_date", recv_date)

            if not price:
                print(f"    ⚠️  Kein Preis/NAV gefunden")
                continue

            # Alle (Datum, NAV)-Paare aus dem Excel — ein Email kann mehrere Stichtage enthalten
            # (z.B. 01.07-Email: "Net Asset Value by 01.07.2026" UND "by 30.06.2026")
            nav_data_points = blatt_data.get("nav_data_points", [])

            if nav_data_points:
                # Für jedes Datum-NAV-Paar einen separaten nav_history-Eintrag anlegen.
                # perf_ytd auf alle Punkte schreiben — wird später bei Duplikaten gemergt.
                pt_ytd = round(float(perf_ytd), 4) if perf_ytd is not None else None
                for point in nav_data_points:
                    pt_nav  = point["nav"]
                    pt_date = point["date"]
                    print(f"    ✅ NAV-Paar: {pt_date} = {pt_nav/1e6:.2f} Mio. € | YTD={pt_ytd}")
                    new_entries.append((fid, {
                        "date": pt_date,
                        "price": round(float(price), 4),
                        "nav": round(float(pt_nav), 2),
                        "perf_ytd": pt_ytd,
                        "source": "measured",
                    }))
            else:
                # Fallback: einzelner Eintrag mit asset_date oder report_date
                if nav is None and price and shares:
                    nav = float(price) * float(shares)
                    print(f"    ℹ️  Nettoverm. berechnet: {price:.4f} × {shares:,.0f} = {nav/1e6:.2f} Mio. €")
                entry_date = asset_date or report_date or recv_date
                print(f"    ✅ {entry_date} | Preis={price:.4f} | NAV={nav/1e6:.2f}M" if nav else f"    ✅ {entry_date} | Preis={price:.4f}")
                new_entries.append((fid, {
                    "date": entry_date,
                    "price": round(float(price), 4),
                    "nav": round(float(nav), 2) if nav else None,
                    "perf_ytd": round(float(perf_ytd), 4) if perf_ytd is not None else None,
                    "source": "measured",
                }))
        except Exception as e:
            print(f"    ❌ Fehler: {e}")
            traceback.print_exc()

    # Bestehende + neue Einträge zusammenführen, deduplizieren.
    # Chronologische Sortierung der new_entries stellt sicher, dass spätere Emails
    # (mit offiziellen Closing-Werten) frühere vorläufige Werte überschreiben.
    for fid, entry in new_entries:
        if fid not in nav_history:
            nav_history[fid] = []
        nav_history[fid].append(entry)

    for fid in nav_history:
        by_date = {}
        for e in sorted(nav_history[fid], key=lambda x: x["date"]):
            existing = by_date.get(e["date"])
            if existing:
                # Merge: neue Werte überschreiben, aber nie einen vorhandenen Wert mit null ersetzen
                merged = {**existing, **e}
                for k in ("nav", "price", "perf_ytd"):
                    if merged.get(k) is None and existing.get(k) is not None:
                        merged[k] = existing[k]
                by_date[e["date"]] = merged
            else:
                by_date[e["date"]] = e
        nav_history[fid] = sorted(by_date.values(), key=lambda x: x["date"])

    total_entries = sum(len(v) for v in nav_history.values())
    print(f"\n✅ NAV-History: {total_entries} Einträge über alle Fonds")
    return nav_history


def backfill_holdings_history(access_token, existing_history):
    """Liest alle historischen INVENTARLISTE-Mails und befüllt holdings_history."""
    messages = find_all_historical_emails(access_token)
    holdings_history = {k: dict(v) for k, v in existing_history.items()}  # deep copy

    # Gruppiere Mails nach Fund + Datum
    mail_map = {}  # {(fid, date_str): msg_id}
    for msg in messages:
        subj   = msg.get("subject", "")
        subj_up = subj.upper()
        if "INVENTARLISTE" not in subj_up:
            continue
        recv_date = msg.get("receivedDateTime", "")[:10]  # YYYY-MM-DD
        for fund in FUNDS:
            fid = fund["id"]
            if fid not in subj:
                continue
            key = (fid, recv_date)
            if key not in mail_map:
                mail_map[key] = msg["id"]

    total = len(mail_map)
    print(f"\n📋 {total} historische INVENTARLISTE-Mails gefunden")

    done = 0
    for (fid, recv_date), msg_id in sorted(mail_map.items(), key=lambda x: x[0][1]):
        # Überspringe bereits vorhandene Snapshots
        if fid in holdings_history and recv_date in holdings_history[fid]:
            done += 1
            print(f"  ♻️  {fid} {recv_date} bereits vorhanden, überspringe")
            continue

        done += 1
        print(f"  [{done}/{total}] Lade {fid} {recv_date}…")
        try:
            xlsx_bytes, filename = download_attachment(access_token, msg_id, ".xlsx")
            if not xlsx_bytes:
                print(f"    ⚠️  Kein Anhang")
                continue
            parsed = parse_excel(xlsx_bytes, fid)
            holdings = parsed.get("holdings", [])
            if not holdings:
                print(f"    ⚠️  Keine Holdings geparst")
                continue
            snap = [
                {"isin": h.get("isin",""), "name": h.get("name",""),
                 "qty": h.get("qty"), "mv_eur": h.get("mv_eur"), "weight": h.get("weight")}
                for h in holdings if h.get("isin") and h["isin"] not in ("None","")
            ]
            if fid not in holdings_history:
                holdings_history[fid] = {}
            holdings_history[fid][recv_date] = snap
            print(f"    ✅ {len(snap)} Positionen gespeichert")
        except Exception as e:
            print(f"    ❌ Fehler: {e}")

    return holdings_history


def find_latest_emails(access_token):
    """Holt die neuesten Mails je Fond — INVENTARBLATT (NAV) + INVENTARLISTE (Holdings)."""
    path = (
        "/me/messages"
        "?$top=100"
        "&$orderby=receivedDateTime desc"
        "&$select=id,subject,receivedDateTime,hasAttachments,from"
    )
    data = graph_get(access_token, path)
    messages = data.get("value", [])
    print(f"📧 {len(messages)} Mails geladen, filtere nach {SENDER_EMAIL}")

    # Pro Fund beide Mail-Typen separat merken
    fund_mails = {}  # {fid: {"blatt": msg, "liste": msg}}
    for msg in messages:
        sender = msg.get("from", {}).get("emailAddress", {}).get("address", "").lower()
        subj   = msg.get("subject", "")
        if sender != SENDER_EMAIL.lower():
            continue
        subj_up = subj.upper()
        for fund in FUNDS:
            fid = fund["id"]
            if fid not in subj:
                continue
            if fid not in fund_mails:
                fund_mails[fid] = {"blatt": None, "liste": None}
            if "INVENTARBLATT" in subj_up and fund_mails[fid]["blatt"] is None:
                fund_mails[fid]["blatt"] = msg
                print(f"  📋 BLATT {fid}: {subj[:55]} ({msg['receivedDateTime'][:10]})")
            elif "INVENTARLISTE" in subj_up and fund_mails[fid]["liste"] is None:
                fund_mails[fid]["liste"] = msg
                print(f"  📊 LISTE {fid}: {subj[:55]} ({msg['receivedDateTime'][:10]})")

    print(f"  Gefunden: { {k: {t: bool(v) for t,v in d.items()} for k,d in fund_mails.items()} }")
    return fund_mails


def download_attachment(access_token, message_id, filename_contains):
    """Lädt den ersten Attachment herunter, dessen Name den String enthält."""
    path = f"/me/messages/{message_id}/attachments"
    data = graph_get(access_token, path)
    attachments = data.get("value", [])

    for att in attachments:
        name = att.get("name", "")
        if filename_contains.lower() in name.lower() or name.lower().endswith(".xlsx"):
            att_id = att["id"]
            print(f"  📎 Lade Anhang: {name}")
            # Binärer Download
            content = graph_get_bytes(access_token, f"/me/messages/{message_id}/attachments/{att_id}/$value")
            return content, name

    # Fallback: contentBytes aus der Metadaten-Antwort
    for att in attachments:
        if att.get("contentBytes"):
            name = att.get("name", "attachment.xlsx")
            print(f"  📎 Lade Anhang (base64): {name}")
            return base64.b64decode(att["contentBytes"]), name

    print(f"  ⚠️  Kein .xlsx Anhang gefunden in Nachricht {message_id}")
    return None, None


# ─── Excel Parser ─────────────────────────────────────────────────────────────
def parse_excel(xlsx_bytes, fund_id):
    """Parst INVENTARBLATT + INVENTARLISTE aus den xlsx-Bytes."""
    try:
        import openpyxl
    except ImportError:
        print("pip install openpyxl")
        sys.exit(1)

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.write(xlsx_bytes)
    tmp.close()

    wb = openpyxl.load_workbook(tmp.name, data_only=True)
    os.unlink(tmp.name)

    sheet_names = wb.sheetnames
    print(f"  📊 Sheets: {sheet_names}")

    result = {}

    # ── INVENTARBLATT ──────────────────────────────────────────────────────
    blatt = None
    for name in sheet_names:
        if "INVENTARBLATT" in name.upper() or "BLATT" in name.upper():
            blatt = wb[name]
            break
    if blatt is None and sheet_names:
        # Versuche erstes Sheet
        blatt = wb[sheet_names[0]]

    if blatt:
        result.update(_parse_inventarblatt(blatt))

    # ── INVENTARLISTE ──────────────────────────────────────────────────────
    for sheet_name in sheet_names:
        sn = sheet_name.upper()
        if "EQUIT" in sn or "AKTIE" in sn:
            result["holdings"] = _parse_positions(wb[sheet_name])
        elif "ACCOUNT" in sn or "KONTO" in sn:
            accs = _parse_positions(wb[sheet_name])
            result.setdefault("holdings", []).extend(accs)
        elif "COUNTRY" in sn or "LAND" in sn:
            result["countries"] = _parse_allocation(wb[sheet_name])
        elif "CURRENCY" in sn or "WÄHR" in sn:
            result["currencies"] = _parse_allocation(wb[sheet_name])
        elif "SEC. TYPE" in sn or "SEKTOR" in sn or "SECTOR" in sn:
            result["sectors"] = _parse_allocation(wb[sheet_name])

    return result


def _parse_inventarblatt(ws):
    """Liest NAV, Preis/Anteil, BVI-Performance aus dem Inventarblatt."""
    data = {}
    rows = list(ws.iter_rows(values_only=True))

    # Debug: alle Zeilen mit Inhalt (max 80)
    print(f"    [BLATT] Sheet '{ws.title}':")
    for i, row in enumerate(rows[:80]):
        non_empty = [(j, str(c)) for j, c in enumerate(row) if c is not None]
        if non_empty:
            print(f"      Row {i+1}: {non_empty}")

    nav_data_points = []  # alle (Datum, NAV)-Paare aus "Net Asset Value by..."-Zeilen

    def _extract_date_from_row(row, row_str):
        """Extrahiert Datum aus einer Zeile — unterstützt DD.MM.YYYY, YYYY-MM-DD und datetime-Objekte."""
        # 1. DD.MM.YYYY in row_str (Zelle als Text)
        m = re.search(r'(\d{1,2})\.(\d{1,2})\.(\d{4})', row_str)
        if m:
            day, mon, yr = m.groups()
            return f"{yr}-{mon.zfill(2)}-{day.zfill(2)}"
        # 2. YYYY-MM-DD in row_str (openpyxl datetime als String)
        m2 = re.search(r'(\d{4})-(\d{2})-(\d{2})', row_str)
        if m2:
            return m2.group(0)[:10]
        # 3. datetime-Objekt direkt in einer Zelle
        for cell in row:
            if isinstance(cell, (datetime, date)) and not isinstance(cell, bool):
                return str(cell)[:10]
        return None

    for i, row in enumerate(rows):
        row_str = " ".join(str(c) for c in row if c is not None).upper()

        # Gesamtvermögen / NAV — viele mögliche Labels.
        # Enthält die Zeile ein Datum (z.B. "Net Asset Value by 30.06.2026"),
        # wird das Datum-NAV-Paar separat gesammelt (für korrekte Monatsend-Zuordnung).
        if any(k in row_str for k in ["GESAMTVERM", "FONDSVERM", "NETTOVERM", "TOTAL NET ASSET",
                                       "TOTAL ASSETS", "FUND VOLUME", "INVENTARWERT GESAMT",
                                       "NET ASSET VALUE", "GESAMT"]):
            nav_val = None
            for cell in row:
                if isinstance(cell, (int, float)) and not isinstance(cell, bool) and cell > 1_000_000:
                    nav_val = float(cell)
                    break
            if nav_val:
                row_date = _extract_date_from_row(row, row_str)
                if row_date:
                    nav_data_points.append({"date": row_date, "nav": nav_val})
                    print(f"    → NAV-Paar Zeile {i+1}: {row_date} = {nav_val/1e6:.2f} Mio. €")
                else:
                    data["nav"] = nav_val  # kein Datum → als generischer NAV speichern
                    print(f"    → NAV Zeile {i+1}: {nav_val/1e6:.2f} Mio. € (kein Datum)")

        # Rücknahmepreis / NAV per share
        # Direkter Match auf "Asset (by DD.MM.YYYY)" — Spalte 2 = Anteile, Spalte 7 = Redemption price
        if "ASSET (BY" in row_str or "ASSET(BY" in row_str:
            # Bewertungsdatum: erste Fundstelle = Vortags-Closing-Datum
            if not data.get("asset_date"):
                d = _extract_date_from_row(row, row_str)
                if d:
                    data["asset_date"] = d
                    print(f"    → Bewertungsdatum (Asset-Zeile {i+1}): {d}")
            # Spalte 2 = Issued/Anteile
            if len(row) > 2 and row[2] is not None:
                try:
                    v = float(row[2])
                    if v > 1000:
                        data["shares"] = v
                        print(f"    → Anteile (Asset-Zeile {i+1}): {v:,.0f}")
                except (TypeError, ValueError):
                    pass
            # Spalte 7 = Redemption price, Spalte 5 = Unit Price
            for col_idx in [7, 6, 5]:
                if col_idx < len(row) and row[col_idx] is not None:
                    try:
                        v = float(row[col_idx])
                        if 10 < v < 5000:
                            data["nav_per_share"] = v
                            print(f"    → Preis (Asset-Zeile {i+1}, col {col_idx}): {v}")
                            break
                    except (TypeError, ValueError):
                        pass
        elif any(k in row_str for k in ["RÜCKNAHME", "ANTEILSWERT", "INVENTARWERT JE",
                                       "REDEMPTION PRICE", "NET ASSET VALUE PER",
                                       "PRICE PER UNIT", "VALUE PER SHARE", "UNIT VALUE",
                                       "ANTEILSPR", "RECHENWERT", "FONDSKURS", "KURS JE",
                                       "PREIS JE", "AUSGABE", "PRICE PER"]):
            for cell in row:
                try:
                    v = float(cell)
                    if 1 < v < 100_000:
                        data["nav_per_share"] = v
                        print(f"    → Preis gefunden Zeile {i+1}: {v}")
                        break
                except (TypeError, ValueError):
                    pass

        # Anzahl Anteile
        if any(k in row_str for k in ["ANTEILE", "UNITS", "SHARES OUTSTANDING", "AUSGEGEBEN"]):
            if any(k in row_str for k in ["UMLAUF", "AUSST", "OUTSTANDING", "ISSUED", "GESAMT"]):
                for cell in row:
                    if isinstance(cell, (int, float)) and cell > 100:
                        data["shares"] = float(cell)
                        print(f"    → Anteile gefunden Zeile {i+1}: {cell}")
                        break

        # BVI Performance
        if "BVI" in row_str or "PERFORMANCE" in row_str or "RENDITE" in row_str:
            if any(k in row_str for k in ["01.01", "JAHRESBEG", "YTD", "YEAR TO DATE", "SEIT 01.01"]):
                for cell in row:
                    if isinstance(cell, (int, float)) and -50 < cell < 200:
                        data["perf_ytd"] = float(cell)
                        print(f"    → YTD gefunden Zeile {i+1}: {cell}")
                        break
            if any(k in row_str for k in ["01.10", "GESCHÄFTSJ", "GJ", "FISCAL", "FISKAL", "SEIT 01.10"]):
                for cell in row:
                    if isinstance(cell, (int, float)) and -50 < cell < 200:
                        data["perf_fy"] = float(cell)
                        print(f"    → FY gefunden Zeile {i+1}: {cell}")
                        break

    # NAV-Daten-Paare auswerten:
    # Falls mehrere (Datum, NAV)-Paare gefunden (z.B. "01.07" und "30.06"),
    # alle speichern. Der neueste Eintrag wird als Primär-NAV verwendet (für tagesaktuelle Anzeige).
    # Das ältere Datum (= Monatsultimo) wird als asset_date gesetzt falls noch nicht vorhanden.
    if nav_data_points:
        sorted_points = sorted(nav_data_points, key=lambda x: x["date"])
        data["nav_data_points"] = sorted_points
        # Primär-NAV = neuester Wert (für tagesaktuelle Verwendung im Full-Run)
        data["nav"] = sorted_points[-1]["nav"]
        # asset_date = ältester Wert (= Vortags-Closing, für korrekten Backfill-Eintrag)
        if not data.get("asset_date"):
            data["asset_date"] = sorted_points[0]["date"]
        print(f"    → NAV-Paare gesamt: {[(p['date'], round(p['nav']/1e6,2)) for p in sorted_points]}")
        print(f"    → asset_date={data.get('asset_date')}, NAV(primär)={data['nav']/1e6:.2f} Mio. €")

    # Datum aus Zellen (für report_date)
    for row in rows[:10]:
        for cell in row:
            if isinstance(cell, (datetime, date)):
                data["report_date"] = str(cell)[:10]
                break

    # Fallback Rücknahmepreis: Suche Zeilen 30–65 nach plausibler Zahl (10–5000 EUR)
    # Plausibilitätsprüfung: NAV / Preis muss > 1000 Anteile ergeben
    if not data.get("nav_per_share") and data.get("nav"):
        nav_val = data["nav"]
        for i, row in enumerate(rows[29:65], start=30):
            for cell in row:
                try:
                    cell_f = float(cell)
                except (TypeError, ValueError):
                    continue
                if 10 < cell_f < 5000:
                    implied_shares = nav_val / cell_f
                    if implied_shares > 1000:
                        data["nav_per_share"] = cell_f
                        print(f"    → Preis (Fallback Zeile {i}): {cell_f} → {implied_shares:,.0f} Anteile impl.")
                        break
            if data.get("nav_per_share"):
                break

    print(f"    [BLATT] Ergebnis: {data}")
    return data


def _parse_positions(ws):
    """Liest Holdings-Positionen."""
    rows = list(ws.iter_rows(values_only=True))
    holdings = []

    # Header-Zeile finden
    header_idx = None
    for i, row in enumerate(rows):
        row_str = " ".join(str(c) for c in row if c is not None).upper()
        if ("ISIN" in row_str or "WKN" in row_str) and ("NAME" in row_str or "BEZEICH" in row_str):
            header_idx = i
            break

    if header_idx is None:
        # Fallback: ab Zeile 3 lesen
        header_idx = 2

    headers = [str(c).strip() if c else "" for c in rows[header_idx]]
    print(f"    [LISTE] Header-Zeile {header_idx+1}: {[(j,h) for j,h in enumerate(headers) if h]}")

    # Spalten-Indices bestimmen
    col_map = {}
    for j, h in enumerate(headers):
        hu = h.upper()
        if "ISIN" in hu:                       col_map["isin"] = j
        elif "NAME" in hu or "BEZEICH" in hu:  col_map["name"] = j
        elif "COUNTRY" in hu or "LAND" in hu:  col_map["country"] = j
        elif "SECTOR" in hu or "BRANCHE" in hu or "SEC.TYPE" in hu: col_map["sector"] = j
        elif "CURRENCY" in hu or "WÄHRUNG" in hu: col_map["currency"] = j
        # Marktwert in Fondswährung (EUR)
        elif "MKT VAL" in hu and ("EUR" in hu or "FNDCCY" in hu or "FUND" in hu):
            col_map["mv_eur"] = j
        elif any(k in hu for k in [
                "P&L", "P/L", "G&V", "G/V", "GEWINN", "UNREALIZED", "UNREAL",
                "GAIN/LOSS", "GAIN LOSS", "BOOK PROFIT", "BUCHGEWINN", "BUCHGEWINN/-VERLUST",
                "BUCHGEWINN/-V", "PROFIT", "VERLUST", "BWG", "BW-GEWINN",
                "UNREALISED", "UNREALIZ", "GAIN", "ACCRU"]):
            col_map.setdefault("pl", j)
        elif "WEIGHT" in hu or "ANTEIL" in hu or "%" in hu:
            col_map.setdefault("weight", j)
        elif "COST" in hu or "EINSTAND" in hu or "KAUFPREIS" in hu or "BOOK VALUE" in hu or "BUCHWERT" in hu:
            col_map.setdefault("cost", j)
        elif "PRICE" in hu or "KURS" in hu:
            col_map.setdefault("price", j)
        elif any(k in hu for k in ["QUANTITY", "STÜCK", "STUCKZAHL", "NOMINAL", "QTY",
                                    "ANZAHL", "BESTAND", "ANTEILE", "NENN", "UNITS",
                                    "SHARES", "VOLUME", "VOLUMEN", "AMOUNT"]):
            col_map.setdefault("qty", j)

    print(f"    [LISTE] col_map: {col_map}")

    # Fallback für Marktwert: Spalte 15 (aus Analyse der echten Daten)
    if "mv_eur" not in col_map:
        col_map["mv_eur"] = 15
        print(f"    [LISTE] mv_eur Fallback: Spalte 15")

    # Fallback für P&L: Spalte nach mv_eur suchen die pos+neg Werte hat
    if "pl" not in col_map and "mv_eur" in col_map:
        mv_idx = col_map["mv_eur"]
        data_rows = [r for r in rows[header_idx + 1:] if any(r)][:30]
        for try_idx in range(mv_idx + 1, min(mv_idx + 8, len(headers))):
            vals = []
            for r in data_rows:
                if try_idx < len(r) and r[try_idx] is not None:
                    try:
                        v = float(r[try_idx])
                        vals.append(v)
                    except (TypeError, ValueError):
                        pass
            has_pos = any(v > 0 for v in vals)
            has_neg = any(v < 0 for v in vals)
            if has_pos and has_neg and len(vals) >= 3:
                col_map["pl"] = try_idx
                print(f"    [LISTE] P&L Fallback: Spalte {try_idx} (pos+neg Werte)")
                break

    for row in rows[header_idx + 1:]:
        if not any(row):
            continue
        # Name-Spalte muss gefüllt sein
        name_val = row[col_map.get("name", 1)] if len(row) > col_map.get("name", 1) else None
        if not name_val or str(name_val).strip() in ("", "None", "Total", "Gesamt"):
            continue

        def get_col(key, default=None):
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return default
            v = row[idx]
            if v is None:
                return default
            return v

        mv_raw = get_col("mv_eur")
        try:
            mv = float(mv_raw) if mv_raw is not None else 0.0
        except (TypeError, ValueError):
            mv = 0.0

        pl_raw = get_col("pl")
        try:
            pl = float(pl_raw) if pl_raw is not None else None
        except (TypeError, ValueError):
            pl = None

        w_raw = get_col("weight")
        try:
            w = float(w_raw) if w_raw is not None else None
        except (TypeError, ValueError):
            w = None

        cost_raw = get_col("cost")
        try:
            cost = float(cost_raw) if cost_raw is not None else None
        except (TypeError, ValueError):
            cost = None

        price_raw = get_col("price")
        try:
            price = float(price_raw) if price_raw is not None else None
        except (TypeError, ValueError):
            price = None

        h = {
            "isin":     str(get_col("isin", "")),
            "name":     str(name_val).strip(),
            "country":  str(get_col("country", "Unbekannt")).strip(),
            "sector":   str(get_col("sector", "Sonstiges")).strip(),
            "currency": str(get_col("currency", "EUR")).strip(),
            "mv_eur":   mv,
            "pl":       pl,
            "weight":   w,
            "cost":     cost,
            "price":    price,
        }
        if mv != 0.0 or pl is not None:
            holdings.append(h)

    return holdings


def _parse_allocation(ws):
    """Liest Allokations-Tabellen (Country / Currency / Sector).
    Versucht Prozent-Spalte zu finden; fällt auf erste numerische Spalte zurück.
    """
    rows = list(ws.iter_rows(values_only=True))

    # Header-Zeile finden und Prozent-Spalte identifizieren
    pct_col = None
    data_start = 0
    for i, row in enumerate(rows[:10]):
        row_str = " ".join(str(c) for c in row if c is not None).upper()
        if any(k in row_str for k in ["%", "WEIGHT", "ANTEIL", "PERCENT", "GEWICHT"]):
            # Finde die Prozent-Spalte
            for j, cell in enumerate(row):
                if cell and any(k in str(cell).upper() for k in ["%", "WEIGHT", "ANTEIL", "PERCENT", "GEWICHT"]):
                    pct_col = j
                    break
            data_start = i + 1
            break

    print(f"    [ALLOC] Sheet '{ws.title}': pct_col={pct_col}, data_start={data_start}")

    result = []
    skip_labels = {"", "None", "Total", "Gesamt", "Land", "Country", "Sektor", "Sector",
                   "Währung", "Currency", "Name", "Bezeichnung"}
    for row in rows[data_start:]:
        if len(row) < 2:
            continue
        label = row[0]
        if not label or str(label).strip() in skip_labels:
            continue
        label_str = str(label).strip()
        # Abbruch bei Summenzeilen
        if any(k in label_str.upper() for k in ["TOTAL", "SUMME", "GESAMT", "SUM"]):
            continue

        val = None
        if pct_col is not None and pct_col < len(row):
            cell = row[pct_col]
            if isinstance(cell, (int, float)) and cell != 0:
                val = float(cell)

        if val is None:
            # Fallback: bevorzuge kleine Zahlen (0-100) als %, vermeide Millionenbeträge
            for cell in row[1:]:
                if isinstance(cell, (int, float)) and cell != 0:
                    if 0 < abs(cell) <= 100:
                        val = float(cell)
                        break
            if val is None:
                # letzter Fallback: irgendeine Zahl
                for cell in row[1:]:
                    if isinstance(cell, (int, float)) and cell != 0:
                        val = float(cell)
                        break

        if val is not None:
            result.append({"label": label_str, "value": val})

    print(f"    [ALLOC] {len(result)} Einträge: {result[:5]}")
    return result


# ─── Kennzahlen berechnen ─────────────────────────────────────────────────────
def compute_kpis(fund_data):
    holdings = fund_data.get("holdings", [])
    nav = fund_data.get("nav", 0)

    total_mv   = sum(h["mv_eur"] for h in holdings)
    total_pl   = sum(h["pl"] for h in holdings if h["pl"] is not None)
    pos_pl     = sum(h["pl"] for h in holdings if h["pl"] and h["pl"] > 0)
    neg_pl     = sum(h["pl"] for h in holdings if h["pl"] and h["pl"] < 0)
    equities_mv = total_mv

    # HHI (Herfindahl-Hirschman Index) – basierend auf Gewichtung
    weights = []
    for h in holdings:
        w = h.get("weight")
        if w is None and nav and nav > 0:
            w = h["mv_eur"] / nav * 100
        if w:
            weights.append(w / 100)
    hhi = sum(w**2 for w in weights) * 10000 if weights else 0

    # Top 10 nach Gewicht
    sorted_h = sorted(holdings, key=lambda x: x["mv_eur"], reverse=True)
    top10_mv  = sum(h["mv_eur"] for h in sorted_h[:10])
    top10_pct = top10_mv / nav * 100 if nav else 0

    # Gewinn-/Verlust-Positionen
    win_positions  = len([h for h in holdings if h.get("pl") and h["pl"] > 0])
    total_with_pl  = len([h for h in holdings if h.get("pl") is not None])
    win_rate = win_positions / total_with_pl * 100 if total_with_pl else 0

    fund_data.update({
        "total_pl":   total_pl,
        "pos_pl":     pos_pl,
        "neg_pl":     neg_pl,
        "equities_mv": equities_mv,
        "hhi":        round(hhi, 1),
        "top10_weight": round(top10_pct, 1),
        "win_rate":   round(win_rate, 1),
    })

    # Länder- und Sektor-Allokation IMMER aus Holdings berechnen (zuverlässiger als dedizierte Sheets)
    if holdings:
        total_mv_h = sum(h["mv_eur"] for h in holdings if h["mv_eur"])
        if total_mv_h > 0:
            ctry_mv = {}
            for h in holdings:
                c = h.get("country") or "Unbekannt"
                if c not in ("None", "", "Unbekannt"):
                    ctry_mv[c] = ctry_mv.get(c, 0) + (h["mv_eur"] or 0)
            if ctry_mv:
                fund_data["countries"] = [
                    {"label": k, "value": round(v / total_mv_h * 100, 2)}
                    for k, v in sorted(ctry_mv.items(), key=lambda x: x[1], reverse=True)
                    if v > 0
                ]
                print(f"    [KPI] Länder aus Holdings: {len(fund_data['countries'])} Einträge")

            sec_mv = {}
            for h in holdings:
                s = h.get("sector") or "Sonstiges"
                if s not in ("None", ""):
                    sec_mv[s] = sec_mv.get(s, 0) + (h["mv_eur"] or 0)
            if sec_mv:
                fund_data["sectors"] = [
                    {"label": k, "value": round(v / total_mv_h * 100, 2)}
                    for k, v in sorted(sec_mv.items(), key=lambda x: x[1], reverse=True)
                    if v > 0
                ]
                print(f"    [KPI] Sektoren aus Holdings: {len(fund_data['sectors'])} Einträge")

    return fund_data


# ─── Änderungen erkennen ──────────────────────────────────────────────────────
def detect_changes(current_holdings, prev_holdings):
    """Vergleicht aktuelle vs. gestrige Holdings."""
    if not prev_holdings:
        return {"added": [], "removed": [], "increased": [], "decreased": [], "date_prev": None}

    curr_map = {h["isin"]: h for h in current_holdings if h.get("isin") and h["isin"] != "None"}
    prev_map = {h["isin"]: h for h in prev_holdings  if h.get("isin") and h["isin"] != "None"}

    added   = [curr_map[i] for i in curr_map if i not in prev_map]
    removed = [prev_map[i] for i in prev_map if i not in curr_map]

    increased = []
    decreased = []
    for isin in curr_map:
        if isin in prev_map:
            # Stückzahl vergleichen (Kauf/Verkauf durch Fondsmanager)
            c_qty = curr_map[isin].get("qty") or 0
            p_qty = prev_map[isin].get("qty") or 0
            if c_qty and p_qty and abs(c_qty - p_qty) / max(abs(p_qty), 1) > 0.005:
                diff_pct = (c_qty - p_qty) / abs(p_qty) * 100
                entry = {**curr_map[isin], "change_pct": round(diff_pct, 2),
                         "prev_qty": p_qty, "curr_qty": c_qty}
                if c_qty > p_qty:
                    increased.append(entry)
                else:
                    decreased.append(entry)

    increased.sort(key=lambda x: abs(x["change_pct"]), reverse=True)
    decreased.sort(key=lambda x: abs(x["change_pct"]), reverse=True)

    return {
        "added":     added[:20],
        "removed":   removed[:20],
        "increased": increased[:10],
        "decreased": decreased[:10],
        "date_prev": None,  # wird vom Aufrufer gesetzt
    }


# ─── Preishistorie (BVI back-calculation) ────────────────────────────────────
def build_price_history(nav_per_share, perf_ytd, perf_fy, nav_per_share_prev=None):
    today = date.today()
    hist = []

    # GJ-Start (01.10. des Vorjahres)
    fy_start_year = today.year - 1 if today.month < 10 else today.year
    fy_start = date(fy_start_year, 10, 1)
    if perf_fy is not None and nav_per_share:
        fy_price = nav_per_share / (1 + perf_fy / 100)
        hist.append({
            "date":  fy_start.isoformat(),
            "label": f"01.10.{fy_start_year}",
            "price": round(fy_price, 4),
            "note":  "GJ-Start (berechnet aus BVI)",
        })

    # Jahresstart (01.01.)
    ytd_start = date(today.year, 1, 1)
    if perf_ytd is not None and nav_per_share:
        ytd_price = nav_per_share / (1 + perf_ytd / 100)
        hist.append({
            "date":  ytd_start.isoformat(),
            "label": f"01.01.{today.year}",
            "price": round(ytd_price, 4),
            "note":  "Jahresstart (berechnet aus BVI)",
        })

    # Vortag (aus prev_data)
    if nav_per_share_prev is not None:
        from datetime import timedelta
        prev_date = today - timedelta(days=1)
        # Wochenenden überspringen
        while prev_date.weekday() >= 5:
            prev_date -= timedelta(days=1)
        hist.append({
            "date":  prev_date.isoformat(),
            "label": prev_date.strftime("%d.%m.%Y"),
            "price": round(nav_per_share_prev, 4),
            "note":  "Vortag (Inventarblatt)",
        })

    # Heute
    if nav_per_share:
        hist.append({
            "date":  today.isoformat(),
            "label": today.strftime("%d.%m.%Y"),
            "price": round(nav_per_share, 4),
            "note":  "Heute (Inventarblatt)",
        })

    return hist


# ─── News fetchen ─────────────────────────────────────────────────────────────
def summarize_news(company_name, articles, anthropic_key):
    """Fasst News-Schlagzeilen via Claude Haiku zu Überschrift + Fließtext zusammen."""
    if not anthropic_key or not articles:
        return None
    articles_text = ""
    for a in articles[:5]:  # max 5 Artikel pro Unternehmen
        title = (a.get("title") or "")[:120]
        desc  = (a.get("desc") or "")[:150]
        content = f"{title} — {desc}" if desc and title not in desc else title
        src = f" ({a['source']})" if a.get("source") else ""
        articles_text += f"- {content}{src}\n"

    prompt = (
        f"Du fasst Nachrichtenartikel für Investoren zusammen.\n"
        f"Unternehmen: {company_name}\n\n"
        f"Artikel:\n{articles_text}\n\n"
        f"STRENGE REGELN – lies sie sorgfältig:\n"
        f"1. Verwende AUSSCHLIESSLICH Informationen die WÖRTLICH im obigen Text stehen.\n"
        f"2. VERBOTEN ohne expliziten Beleg im Text:\n"
        f"   - Indexänderungen (ATX, DAX, MSCI etc. Aufnahme oder Ausschluss)\n"
        f"   - Übernahmen, Fusionen, M&A\n"
        f"   - CEO- oder Managementwechsel\n"
        f"   - Konkrete Zahlen (Umsatz, Gewinn, Kurs) die nicht im Text stehen\n"
        f"   - Zukunftsprognosen die nicht zitiert werden\n"
        f"3. Im Zweifel: weglassen oder IRRELEVANT antworten.\n"
        f"4. Wenn kein Artikel wirklich zu {company_name} passt (Geschäft, Zahlen, Strategie): antworte nur mit IRRELEVANT\n\n"
        f"Antworte in diesem Format:\n"
        f"HEADLINE: [präzise Überschrift, nur was belegt ist]\n"
        f"TEXT: [2–3 Sätze, direkt, nur belegte Fakten – lieber kürzer als spekulativ]"
    )
    body = json.dumps({
        "model": "claude-haiku-4-5-20251001",
        "max_tokens": 200,
        "messages": [{"role": "user", "content": prompt}],
    }, ensure_ascii=False).encode("utf-8")
    req = Request(
        "https://api.anthropic.com/v1/messages", data=body, method="POST",
        headers={
            "x-api-key": anthropic_key,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json; charset=utf-8",
        },
    )
    try:
        with urlopen(req, timeout=20) as resp:
            result = json.loads(resp.read())
            raw = result["content"][0]["text"].strip()
        if raw.startswith("IRRELEVANT"):
            return None
        headline, text = "", ""
        for line in raw.splitlines():
            if line.startswith("HEADLINE:"):
                headline = line[9:].strip()
            elif line.startswith("TEXT:"):
                text = line[5:].strip()
        return {"headline": headline, "text": text} if (headline or text) else None
    except HTTPError as e:
        body_err = ""
        try: body_err = e.read().decode("utf-8", errors="replace")[:300]
        except Exception: pass
        print(f"    ⚠️  Haiku-Fehler für {company_name}: HTTP {e.code} – {body_err}")
        return None
    except Exception as e:
        print(f"    ⚠️  Haiku-Fehler für {company_name}: {e}")
        return None


def _clean_news_name(name):
    """Bereinigt Firmennamen für Google News Suche."""
    # Nur echte Finanz-Suffixe entfernen – GROUP/GRP/HOLDINGS bleiben als Namensteil
    clean = re.sub(
        r'\s+(PLC|AG|SE|INC\.?|CORP\.?|LTD\.?|S\.A\.|SA|NV|BV|SPA|CO\.?|'
        r'INH\.?|O\.N\.|BK|BNK|ST\.AKT\.ON|DL-?[\d,\.]+|EO[\s\-]?[\d,\.]+|'
        r'LS-?[\d,\.]+|SF\s*[\d,\.]+|CL\.[A-Z])(\s.*)?$',
        '', name.strip(), flags=re.IGNORECASE
    ).strip()
    return ' '.join(clean.split()[:4])


    # Domains die keine Finanz-News liefern → werden herausgefiltert
_BLOCKLIST_DOMAINS = {
    "xboxdynasty", "filmstarts", "photografix", "connect.de", "computerbild",
    "chip.de", "heise.de", "golem.de", "computerbase", "ign.com", "gamestar",
    "gamesradar", "eurogamer", "kotaku", "polygon.com", "pcgamer", "hardwareluxx",
    "notebookcheck", "techradar", "theverge", "engadget", "wired.com",
    "solidbau", "immobilien", "realestate", "architekt", "bau.de",
}

def _is_finance_relevant(title, source):
    """Prüft ob ein Artikel finanziell relevant ist."""
    src_low = source.lower()
    if any(bl in src_low for bl in _BLOCKLIST_DOMAINS):
        return False
    # Titel-Filter: mindestens ein Finanz-Keyword oder kein offensichtliches Off-Topic
    off_topic = ["rezept", "urlaub", "reise", "gaming", "spiel ", "film ", "serie ",
                 "musik", "mode ", "beauty", "gesundheit", "sport ", "fußball",
                 "küche", "wohnen", "garten"]
    title_low = title.lower()
    if any(kw in title_low for kw in off_topic):
        return False
    return True


def fetch_all_news(companies, max_per_company=8, request_timeout=5, anthropic_key=None, max_summaries=80, prev_news_data=None, max_wall_seconds=300):
    """Fetcht Finanz-News via Google News RSS für alle Unternehmen, optional mit Haiku-Zusammenfassung."""
    import xml.etree.ElementTree as ET
    import time as _time
    from urllib.parse import quote as _quote

    news_data = {}
    items_list = list(companies.items())
    total = len(items_list)
    summarize = bool(anthropic_key)
    summary_count = 0
    deadline = _time.monotonic() + max_wall_seconds
    print(f"\n📰 Fetche News für {total} Unternehmen{f' (KI-Summary für Top {max_summaries})' if summarize else ''} (max {max_wall_seconds}s)…")

    for i, (key, co) in enumerate(items_list):
        if _time.monotonic() > deadline:
            print(f"  ⏱️  Zeitlimit erreicht nach {i} Unternehmen – stoppe News-Fetch.")
            break
        clean = _clean_news_name(co["name"])
        if not clean or len(clean) < 3:
            continue
        q = _quote(clean[:50])
        url = f"https://news.google.com/rss/search?q={q}&hl=de&gl=AT&ceid=AT:de"
        try:
            req = Request(url, headers={
                "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36"
            })
            with urlopen(req, timeout=request_timeout) as resp:
                content = resp.read()
            root = ET.fromstring(content)
            arts = []
            for el in root.findall(".//item"):
                if len(arts) >= max_per_company:
                    break
                title = re.sub(r"<[^>]+>|<!\[CDATA\[|\]\]>", "", el.findtext("title") or "").strip()
                link  = (el.findtext("link") or "#").strip()
                pub   = (el.findtext("pubDate") or "").strip()
                src   = getattr(el.find("source"), "text", "") or ""
                desc  = re.sub(r"<[^>]+>|<!\[CDATA\[|\]\]>", "", el.findtext("description") or "").strip()
                if title and _is_finance_relevant(title, src):
                    arts.append({"title": title, "link": link, "pubDate": pub, "source": src, "desc": desc})
            if arts:
                # Article-Caching: Summary wiederverwenden wenn Top-Artikel unverändert
                prev = (prev_news_data or {}).get(key, {})
                prev_top = ((prev.get("articles") or [{}])[0]).get("title", "")
                curr_top = arts[0].get("title", "")
                if prev_top and curr_top == prev_top and prev.get("summary"):
                    summary = prev["summary"]
                    print(f"    ♻️  Cache für {co['name'][:35]}")
                else:
                    do_summary = summarize and summary_count < max_summaries
                    summary = None
                    if do_summary:
                        _time.sleep(1.5)  # Rate-limit
                        summary = summarize_news(co["name"], arts, anthropic_key)
                    if summary:
                        summary_count += 1
                news_data[key] = {
                    "company": co["name"],
                    "funds": co["funds"],
                    "articles": arts,
                    "summary": summary,
                }
        except Exception as _ne:
            if i < 3 or (i + 1) % 20 == 0:
                print(f"  ⚠️  News-Fehler {co['name'][:30]}: {type(_ne).__name__}: {_ne}")
        if (i + 1) % 20 == 0:
            print(f"  {i+1}/{total}…")
        _time.sleep(0.2)

    print(f"  ✅ {len(news_data)}/{total} Unternehmen mit Artikeln")
    return news_data


# ─── Dashboard HTML generieren ────────────────────────────────────────────────
def generate_html(funds_data, updated_at, nav_history=None, news_data=None, run_log=None, changes_history=None):
    '''Generiert das vollstaendige Dashboard-HTML.'''
    data_json = json.dumps(funds_data, ensure_ascii=False, separators=(',',':'))
    nav_history_json = json.dumps(nav_history or {}, ensure_ascii=False, separators=(',',':'))
    news_data_json = json.dumps(news_data or {}, ensure_ascii=False, separators=(',',':'))
    changes_history_json = json.dumps(changes_history or {}, ensure_ascii=False, separators=(',',':'))
    run_log_json = json.dumps(run_log or [], ensure_ascii=False, separators=(',',':'))
    html = f'''<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Sunrise Dashboard</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/canvas-confetti@1.6.0/dist/confetti.browser.min.js"></script>
<style>
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:system-ui,-apple-system,sans-serif;background:#f5f5f3;color:#1a1a1a;min-height:100vh;font-size:14px}}
a{{color:inherit;text-decoration:none}}
.header{{background:#fff;border-bottom:1px solid #e8e8e6;padding:0 24px;height:56px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:100}}
.header-title{{font-size:16px;font-weight:700;letter-spacing:-0.3px}}
.header-updated{{font-size:12px;color:#888}}
.summary-bar{{background:#f0f0ee;border-bottom:1px solid #e8e8e6;padding:0 24px;display:flex;gap:32px;align-items:center;height:52px}}
.summary-tile{{display:flex;flex-direction:column}}
.summary-label{{font-size:10px;text-transform:uppercase;letter-spacing:0.5px;color:#888;font-weight:600}}
.summary-value{{font-size:15px;font-weight:700;color:#1a1a1a;letter-spacing:-0.3px}}
.summary-value.pos{{color:#1a7c4a}}
.summary-value.neg{{color:#c0392b}}
.fund-selector-wrap{{padding:20px 24px 0}}
.fund-selector{{display:flex;gap:8px;flex-wrap:wrap}}
.pill{{padding:8px 18px;border-radius:100px;border:1.5px solid #d8d8d6;background:#fff;cursor:pointer;font-size:13px;font-weight:500;color:#555;transition:all .15s}}
.pill:hover{{border-color:#aaa;background:#f8f8f6}}
.pill.active{{border-color:var(--fund-color,#3B7DD8);background:var(--fund-color,#3B7DD8);color:#fff}}
.fund-header{{padding:16px 24px 0}}
.fund-name{{font-size:24px;font-weight:700;letter-spacing:-0.5px}}
.fund-isin{{font-size:12px;color:#888;margin-top:2px;font-family:monospace}}
.fund-tags{{display:flex;gap:6px;margin-top:10px;flex-wrap:wrap}}
.fund-tag{{padding:3px 10px;border-radius:4px;background:#f0f0ee;font-size:11px;color:#555;font-weight:500}}
.metrics-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:12px;padding:20px 24px 0}}
.metric-tile{{background:#fff;border-radius:12px;padding:16px 20px;border:1px solid #e8e8e6}}
.metric-label{{font-size:11px;text-transform:uppercase;letter-spacing:0.5px;color:#888;font-weight:600;margin-bottom:8px}}
.metric-value{{font-size:20px;font-weight:700;color:#1a1a1a;letter-spacing:-0.5px}}
.metric-value.pos{{color:#1a7c4a}}
.metric-value.neg{{color:#c0392b}}
.tabs-wrap{{padding:20px 24px 0}}
.tab-nav{{display:flex;border-bottom:2px solid #e8e8e6;overflow-x:auto;scrollbar-width:none}}
.tab-nav::-webkit-scrollbar{{display:none}}
.tab-btn{{padding:10px 18px;font-size:13px;font-weight:500;color:#888;background:none;border:none;cursor:pointer;white-space:nowrap;border-bottom:2px solid transparent;margin-bottom:-2px;transition:all .15s}}
.tab-btn:hover{{color:#1a1a1a}}
.tab-btn.active{{font-weight:600}}
.tab-panel{{display:none;padding:24px 0 40px}}
.tab-panel.active{{display:block}}
.card{{background:#fff;border-radius:12px;padding:20px;border:1px solid #e8e8e6;margin-bottom:20px}}
.card-title{{font-size:13px;font-weight:600;color:#888;margin-bottom:16px}}
.chart-wrap{{height:260px;position:relative}}
.section-title{{font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:0.5px;color:#888;margin-bottom:12px;margin-top:20px}}
.perf-table,.data-table,.monthly-table{{width:100%;border-collapse:collapse;background:#fff;border-radius:12px;overflow:hidden;border:1px solid #e8e8e6}}
.perf-table th,.data-table th,.monthly-table th{{background:#f8f8f6;padding:10px 16px;text-align:left;font-size:11px;text-transform:uppercase;letter-spacing:0.5px;color:#888;font-weight:600;cursor:pointer;user-select:none;white-space:nowrap}}
.perf-table th:hover,.data-table th:hover{{background:#f0f0ee}}
.perf-table td,.data-table td,.monthly-table td{{padding:11px 16px;border-top:1px solid #f0f0ee;font-size:13px;vertical-align:middle}}
td.pos{{color:#1a7c4a;font-weight:600}}
td.neg{{color:#c0392b;font-weight:600}}
td.num{{text-align:right;font-variant-numeric:tabular-nums}}
td.rank{{font-size:12px;color:#aaa;width:36px;text-align:center}}
td.date-cell{{font-size:12px;color:#888;white-space:nowrap}}
.h-name{{font-weight:500}}
.h-isin{{font-size:11px;color:#aaa;margin-top:2px;font-family:monospace}}
.badge{{display:inline-block;padding:3px 9px;border-radius:4px;font-size:11px;font-weight:600}}
.risk-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:16px;margin-top:8px}}
.risk-tile{{background:#fff;border-radius:12px;padding:20px 24px;border:1px solid #e8e8e6}}
.risk-label{{font-size:11px;text-transform:uppercase;letter-spacing:0.5px;color:#888;font-weight:600;margin-bottom:6px}}
.risk-value{{font-size:22px;font-weight:700;color:#1a1a1a;letter-spacing:-0.5px}}
.risk-note{{font-size:11px;color:#aaa;margin-top:4px;line-height:1.4}}
.analyse-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(260px,1fr));gap:16px}}
.donut-legend{{margin-top:12px;display:flex;flex-direction:column;gap:6px}}
.legend-item{{display:flex;align-items:center;gap:8px;font-size:12px}}
.legend-dot{{width:10px;height:10px;border-radius:50%;flex-shrink:0}}
.legend-label{{flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;color:#555}}
.legend-pct{{color:#888;font-variant-numeric:tabular-nums}}
.gv-tabs{{display:flex;gap:8px;margin-bottom:16px}}
.gv-tab{{padding:6px 14px;border-radius:100px;border:1px solid #d8d8d6;background:#fff;cursor:pointer;font-size:12px;font-weight:500;color:#666;transition:all .15s}}
.gv-tab.active{{background:#1a1a1a;border-color:#1a1a1a;color:#fff}}
.gv-grid{{display:grid;grid-template-columns:1fr 1fr;gap:12px}}
.gv-header{{font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:0.5px;padding:6px 10px;border-radius:6px;margin-bottom:8px}}
.gv-header.winners{{background:#dcfce7;color:#1a7c4a}}
.gv-header.losers{{background:#fee2e2;color:#c0392b}}
.gv-row{{display:flex;justify-content:space-between;align-items:center;padding:6px 0;border-bottom:1px solid #f0f0ee;gap:8px}}
.gv-rank{{font-size:11px;color:#aaa;width:20px;flex-shrink:0}}
.gv-name{{font-size:12px;font-weight:500;flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}
.gv-pl{{font-size:12px;font-weight:600;text-align:right;flex-shrink:0}}
.news-grid{{display:grid;gap:16px}}
.news-card{{background:#fff;border-radius:12px;padding:20px;border:1px solid #e8e8e6}}
.news-company{{font-size:15px;font-weight:700;margin-bottom:10px}}
.news-headline{{font-size:14px;font-weight:600;color:#1a1a1a;margin-bottom:6px;line-height:1.4}}
.news-summary{{font-size:12px;color:#666;line-height:1.6;margin-bottom:12px;padding:10px;background:#f8f8f6;border-radius:8px}}
.articles-list{{display:flex;flex-direction:column;gap:8px}}
.article-link{{display:flex;flex-direction:column;gap:2px;padding:8px;border-radius:6px;border:1px solid #f0f0ee;transition:background .15s;color:inherit;text-decoration:none}}
.article-link:hover{{background:#f8f8f6}}
.article-title{{font-size:13px;font-weight:500;line-height:1.4}}
.article-meta{{font-size:11px;color:#aaa}}
.run-log-wrap{{padding:24px 24px 40px;margin-top:8px}}
details{{background:#fff;border-radius:12px;border:1px solid #e8e8e6;overflow:hidden}}
summary{{padding:14px 20px;cursor:pointer;font-size:12px;font-weight:600;color:#888;text-transform:uppercase;letter-spacing:0.5px;list-style:none;display:flex;align-items:center;gap:8px}}
summary::-webkit-details-marker{{display:none}}
summary::before{{content:"\25B6";font-size:10px;transition:transform .15s}}
details[open] summary::before{{transform:rotate(90deg)}}
.empty{{color:#888;font-size:13px;padding:20px 0}}
.news-btn{{padding:7px 16px;border-radius:100px;border:1.5px solid #e8e8e6;background:#fff;cursor:pointer;font-size:13px;font-weight:500;color:#555;transition:all .15s;display:flex;align-items:center;gap:6px}}
.news-btn:hover{{border-color:#3B7DD8;color:#3B7DD8;background:#f0f5ff}}
.news-overlay{{display:none;position:fixed;inset:0;z-index:200;background:rgba(0,0,0,0.45);backdrop-filter:blur(2px)}}
.news-overlay.open{{display:flex;align-items:flex-start;justify-content:flex-end}}
.news-drawer{{background:#f5f5f3;width:min(600px,100vw);height:100vh;overflow-y:auto;box-shadow:-4px 0 24px rgba(0,0,0,0.15);display:flex;flex-direction:column}}
.news-drawer-header{{background:#fff;border-bottom:1px solid #e8e8e6;padding:16px 24px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:10}}
.news-drawer-title{{font-size:16px;font-weight:700}}
.news-drawer-close{{padding:6px 12px;border-radius:8px;border:1px solid #e8e8e6;background:#fff;cursor:pointer;font-size:13px;color:#555}}
.news-drawer-close:hover{{background:#f0f0ee}}
.news-drawer-body{{padding:20px;flex:1}}
</style>
</head>
<body>
<header class="header">
  <span class="header-title">Sunrise Dashboard</span>
  <div style="display:flex;align-items:center;gap:12px">
    <button class="news-btn" id="news-btn" onclick="openNewsOverlay()">&#128240; News</button>
    <span class="header-updated">Stand: {updated_at}</span>
  </div>
</header>
<div class="news-overlay" id="news-overlay" onclick="overlayBgClick(event)">
  <div class="news-drawer">
    <div class="news-drawer-header">
      <span class="news-drawer-title">&#128240; News<span id="news-drawer-title-fund" style="font-weight:400;color:#888"></span></span>
      <button class="news-drawer-close" onclick="closeNewsOverlay()">&#10005; Schlie&#223;en</button>
    </div>
    <div class="news-drawer-body">
      <div id="news-drawer-content" class="news-grid"></div>
    </div>
  </div>
</div>
<div class="summary-bar">
  <div class="summary-tile">
    <span class="summary-label">Gesamt AUM</span>
    <span class="summary-value" id="sum-aum">&#8212;</span>
  </div>
  <div class="summary-tile">
    <span class="summary-label">Gesamt P&amp;L</span>
    <span class="summary-value" id="sum-pl">&#8212;</span>
  </div>
  <div class="summary-tile">
    <span class="summary-label">Stand</span>
    <span class="summary-value">{updated_at}</span>
  </div>
</div>
<div class="fund-selector-wrap">
  <div class="fund-selector" id="fund-selector"></div>
</div>
<div class="fund-header">
  <div class="fund-name" id="fund-name"></div>
  <div class="fund-isin" id="fund-isin"></div>
  <div class="fund-tags">
    <span class="fund-tag">Sonderverm&#246;gen</span>
    <span class="fund-tag">Institutionell</span>
  </div>
</div>
<div class="metrics-grid">
  <div class="metric-tile">
    <div class="metric-label">Nettoverm&#246;gen</div>
    <div class="metric-value" id="metric-nav"></div>
  </div>
  <div class="metric-tile">
    <div class="metric-label">Fondspreis</div>
    <div class="metric-value" id="metric-price"></div>
  </div>
  <div class="metric-tile">
    <div class="metric-label">Performance YTD</div>
    <div class="metric-value" id="metric-ytd"></div>
  </div>
  <div class="metric-tile">
    <div class="metric-label">Positionen</div>
    <div class="metric-value" id="metric-positions"></div>
  </div>
</div>
<div class="tabs-wrap">
  <nav class="tab-nav" id="tab-nav">
    <button class="tab-btn active" data-tab="overview">&#220;bersicht</button>
    <button class="tab-btn" data-tab="holdings">Holdings</button>
    <button class="tab-btn" data-tab="analyse">Analyse</button>
    <button class="tab-btn" data-tab="transactions">Transaktionen</button>
    <button class="tab-btn" data-tab="risk">Risiko</button>
    <button class="tab-btn" data-tab="monthly">Monatsreporting</button>
  </nav>
  <div class="tab-panel active" id="tab-overview">
    <div class="card">
      <div class="card-title">Fondspreis &#8212; letzte 12 Monate</div>
      <div class="chart-wrap"><canvas id="nav-chart"></canvas></div>
    </div>
    <div class="section-title">Performance</div>
    <table class="perf-table">
      <thead><tr><th>Zeitraum</th><th style="text-align:right">Rendite</th></tr></thead>
      <tbody id="perf-table-body"></tbody>
    </table>
    <div class="section-title" style="margin-top:24px">Gewinner &amp; Verlierer</div>
    <div class="card">
      <div class="gv-tabs" id="gv-tabs">
        <button class="gv-tab active" data-gv="total">P&amp;L gesamt</button>
        <button class="gv-tab" data-gv="month">Monat</button>
        <button class="gv-tab" data-gv="day">Tag</button>
      </div>
      <div class="gv-grid" id="gv-container"></div>
    </div>
  </div>
  <div class="tab-panel" id="tab-holdings">
    <table class="data-table" id="holdings-table">
      <thead>
        <tr>
          <th data-col="rank">#</th>
          <th data-col="name">Position</th>
          <th data-col="mv" style="text-align:right">Marktwert</th>
          <th data-col="pct" style="text-align:right">% NAV</th>
          <th data-col="pl" style="text-align:right">P&amp;L</th>
          <th data-col="sector">Sektor</th>
        </tr>
      </thead>
      <tbody id="holdings-body"></tbody>
    </table>
  </div>
  <div class="tab-panel" id="tab-analyse">
    <div class="analyse-grid">
      <div class="card">
        <div class="card-title">Sektoren</div>
        <div class="chart-wrap" style="height:200px"><canvas id="chart-sector"></canvas></div>
        <div class="donut-legend" id="legend-sector"></div>
      </div>
      <div class="card">
        <div class="card-title">L&#228;nder</div>
        <div class="chart-wrap" style="height:200px"><canvas id="chart-country"></canvas></div>
        <div class="donut-legend" id="legend-country"></div>
      </div>
      <div class="card">
        <div class="card-title">W&#228;hrungen</div>
        <div class="chart-wrap" style="height:200px"><canvas id="chart-currency"></canvas></div>
        <div class="donut-legend" id="legend-currency"></div>
      </div>
    </div>
  </div>
  <div class="tab-panel" id="tab-transactions">
    <table class="data-table">
      <thead><tr><th>Datum</th><th>Typ</th><th>Name</th><th style="text-align:right">&#916; %</th><th style="text-align:right">Marktwert</th></tr></thead>
      <tbody id="transactions-body"></tbody>
    </table>
  </div>
  <div class="tab-panel" id="tab-risk">
    <div class="risk-grid">
      <div class="risk-tile">
        <div class="risk-label">Volatilit&#228;t p.a.</div>
        <div class="risk-value" id="risk-vol">&#8212;</div>
        <div class="risk-note">Standardabweichung der t&#228;glichen Log-Renditen, annualisiert mit Faktor &#8730;252.</div>
      </div>
      <div class="risk-tile">
        <div class="risk-label">Max. Drawdown</div>
        <div class="risk-value" id="risk-drawdown">&#8212;</div>
        <div class="risk-note">Gr&#246;&#223;ter Verlust vom H&#246;chststand zum Tiefpunkt (letzte 252 Handelstage).</div>
      </div>
      <div class="risk-tile">
        <div class="risk-label">Sharpe Ratio</div>
        <div class="risk-value" id="risk-sharpe">&#8212;</div>
        <div class="risk-note">Verh&#228;ltnis von annualisierter Rendite zu Volatilit&#228;t (risikoloser Zins: 0 %).</div>
      </div>
    </div>
  </div>
  <div class="tab-panel" id="tab-monthly">
    <table class="monthly-table">
      <thead>
        <tr>
          <th>Monat</th>
          <th style="text-align:right">Nettoverm&#246;gen</th>
          <th style="text-align:right">Fondspreis</th>
          <th style="text-align:right">YTD %</th>
          <th style="text-align:right">&#916; Vormonat</th>
        </tr>
      </thead>
      <tbody id="monthly-body"></tbody>
    </table>
  </div>
</div>
<div class="run-log-wrap">
  <details>
    <summary>Run Log</summary>
    <div style="padding:0 20px 16px">
      <table class="data-table" style="margin-top:12px">
        <thead><tr><th>Zeitstempel</th><th>Status</th><th>Fonds</th><th>Positionen</th><th>News</th><th>Summaries</th></tr></thead>
        <tbody id="run-log-body"></tbody>
      </table>
    </div>
  </details>
</div>
<script>
const FUNDS_DATA={data_json};
const NAV_HISTORY={nav_history_json};
const NEWS_DATA={news_data_json};
const CHANGES_HISTORY={changes_history_json};
const RUN_LOG={run_log_json};
let currentFundIdx=0,navChart=null,chartSector=null,chartCountry=null,chartCurrency=null,holdingsSortCol='pct',holdingsSortDir=-1,gvMode='total';
const fmtEur=new Intl.NumberFormat('de-AT',{{minimumFractionDigits:2,maximumFractionDigits:2}});
const fmtInt=new Intl.NumberFormat('de-AT',{{maximumFractionDigits:0}});
const fundName=n=>(n||'').replace(/^IQAM\s+/i,'');
const fmtNav=n=>n==null?'\u2014':fmtInt.format(Math.round(n))+' \u20ac';
const fmtPrice=n=>n==null?'\u2014':fmtEur.format(n)+' \u20ac';
const fmtPerf=n=>n==null?'\u2014':(n>=0?'+':'')+ fmtEur.format(n)+' %';
const fmtDelta=n=>n==null?'\u2014':(n>=0?'+':'')+ fmtInt.format(Math.round(Math.abs(n)))+' \u20ac';
const colorVal=n=>n==null?'':(n>=0?'color:#1a7c4a':'color:#c0392b');
const esc=s=>(s||'').toString().replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
const DONUT_COLORS=['#3B7DD8','#F4B942','#1a7c4a','#c0392b','#9B59B6','#E67E22','#1ABC9C','#E91E63','#607D8B','#795548','#FF5722','#00BCD4','#8BC34A','#FF9800'];
function renderSummaryBar(){{
  var totalNav=0,totalPl=0;
  FUNDS_DATA.forEach(function(f){{totalNav+=(f.nav||0);totalPl+=(f.total_pl||0);}});
  document.getElementById('sum-aum').textContent=fmtInt.format(Math.round(totalNav))+' \u20ac';
  var el=document.getElementById('sum-pl');
  el.textContent=(totalPl>=0?'+':'')+ fmtInt.format(Math.round(Math.abs(totalPl)))+' \u20ac';
  el.className='summary-value'+(totalPl>=0?' pos':' neg');
}}
function renderFundSelector(){{
  var c=document.getElementById('fund-selector');c.innerHTML='';
  FUNDS_DATA.forEach(function(f,i){{
    var b=document.createElement('button');
    b.className='pill'+(i===currentFundIdx?' active':'');
    b.textContent=fundName(f.name);
    b.style.setProperty('--fund-color',f.color||'#3B7DD8');
    b.onclick=function(){{switchFund(i);}};
    c.appendChild(b);
  }});
}}
function destroyCharts(){{
  if(navChart){{navChart.destroy();navChart=null;}}
  if(chartSector){{chartSector.destroy();chartSector=null;}}
  if(chartCountry){{chartCountry.destroy();chartCountry=null;}}
  if(chartCurrency){{chartCurrency.destroy();chartCurrency=null;}}
}}
function switchFund(idx){{
  currentFundIdx=idx;destroyCharts();renderFundSelector();renderFundHeader();renderMetrics();
  var a=document.querySelector('.tab-btn.active');switchTab(a?a.dataset.tab:'overview');
}}
function renderFundHeader(){{
  var f=FUNDS_DATA[currentFundIdx];
  document.getElementById('fund-name').textContent=fundName(f.name);
  document.getElementById('fund-isin').textContent=f.isin||'';
}}
function renderMetrics(){{
  var f=FUNDS_DATA[currentFundIdx];
  document.getElementById('metric-nav').textContent=fmtNav(f.nav);
  document.getElementById('metric-price').textContent=fmtPrice(f.nav_per_share);
  var ytd=f.perf_ytd,el=document.getElementById('metric-ytd');
  el.textContent=fmtPerf(ytd);
  el.className='metric-value'+(ytd==null?'':(ytd>=0?' pos':' neg'));
  document.getElementById('metric-positions').textContent=(f.holdings||[]).length;
}}
function updateTabColor(){{
  var color=(FUNDS_DATA[currentFundIdx].color)||'#3B7DD8';
  document.querySelectorAll('.tab-btn').forEach(function(b){{
    if(b.classList.contains('active')){{b.style.borderBottomColor=color;b.style.color=color;}}
    else{{b.style.borderBottomColor='';b.style.color='';}}
  }});
}}
function switchTab(tabId){{
  var color=(FUNDS_DATA[currentFundIdx].color)||'#3B7DD8';
  document.querySelectorAll('.tab-btn').forEach(function(b){{
    var a=b.dataset.tab===tabId;b.classList.toggle('active',a);
    b.style.borderBottomColor=a?color:'';b.style.color=a?color:'';
  }});
  document.querySelectorAll('.tab-panel').forEach(function(p){{
    p.classList.toggle('active',p.id==='tab-'+tabId);
  }});
  if(tabId==='overview')renderOverview();
  else if(tabId==='holdings')renderHoldings();
  else if(tabId==='analyse')renderAnalyse();
  else if(tabId==='transactions')renderTransactions();
  else if(tabId==='risk')renderRisk();
  else if(tabId==='monthly')renderMonthly();
  else if(tabId==='news')renderNews();
}}
function getHistory(){{return NAV_HISTORY[FUNDS_DATA[currentFundIdx].id]||[];}}
function calcPerf(history,days){{
  if(!history||history.length<2)return null;
  var latest=history[history.length-1];
  var cutoff=new Date(latest.date);cutoff.setDate(cutoff.getDate()-days);
  var baseline=null;
  for(var i=0;i<history.length;i++){{if(new Date(history[i].date)>=cutoff){{baseline=history[i];break;}}}}
  if(!baseline||baseline.date===latest.date||!baseline.price)return null;
  return((latest.price-baseline.price)/baseline.price)*100;
}}
function calcYTD(history){{
  if(!history||history.length<2)return null;
  var latest=history[history.length-1];
  if(latest.perf_ytd!=null)return latest.perf_ytd;
  var year=new Date(latest.date).getFullYear();
  for(var i=0;i<history.length;i++){{
    if(new Date(history[i].date).getFullYear()===year){{
      var b=history[i];if(!b.price)return null;
      return((latest.price-b.price)/b.price)*100;
    }}
  }}
  return null;
}}
function renderOverview(){{
  var history=getHistory();
  var cutoff=new Date();cutoff.setMonth(cutoff.getMonth()-12);
  var filtered=history.filter(function(h){{return new Date(h.date)>=cutoff;}});
  var f=FUNDS_DATA[currentFundIdx],color=f.color||'#3B7DD8';
  var canvas=document.getElementById('nav-chart');
  if(navChart){{navChart.destroy();navChart=null;}}
  if(filtered.length>0){{
    navChart=new Chart(canvas,{{
      type:'line',
      data:{{
        labels:filtered.map(function(h){{return h.date;}}),
        datasets:[{{
          label:'Fondspreis',
          data:filtered.map(function(h){{return h.price;}}),
          borderColor:color,backgroundColor:color+'20',fill:true,tension:0.3,pointRadius:0,borderWidth:2
        }}]
      }},
      options:{{
        responsive:true,maintainAspectRatio:false,
        plugins:{{
          legend:{{display:false}},
          tooltip:{{mode:'index',intersect:false,callbacks:{{label:function(ctx){{return fmtPrice(ctx.parsed.y);}}}}}}
        }},
        scales:{{
          x:{{grid:{{display:false}},ticks:{{maxTicksLimit:8,maxRotation:0}}}},
          y:{{grid:{{color:'#e8e8e6'}},ticks:{{callback:function(v){{return fmtEur.format(v)+' \u20ac';}}}}}}
        }}
      }}
    }});
  }}
  renderPerfTable(history);renderGV();
}}
function renderPerfTable(history){{
  var periods=[
    {{label:'1 Woche',days:7}},{{label:'1 Monat',days:30}},{{label:'3 Monate',days:90}},
    {{label:'6 Monate',days:180}},{{label:'YTD',ytd:true}},{{label:'1 Jahr',days:365}}
  ];
  var rows='';
  periods.forEach(function(p){{
    var val=p.ytd?calcYTD(history):calcPerf(history,p.days);
    var cls=val==null?'':(val>=0?' pos':' neg');
    rows+='<tr><td>'+p.label+'</td><td class="num'+cls+'">'+( val==null?'\u2014':fmtPerf(val))+'</td></tr>';
  }});
  document.getElementById('perf-table-body').innerHTML=rows;
}}
function renderGV(){{
  var f=FUNDS_DATA[currentFundIdx];
  var allH=(f.holdings||[]);
  var sorted,getVal,fmtVal;
  if(gvMode==='month'){{
    /* Rendite % = pl / Einstandswert (mv_eur - pl) */
    sorted=allH.filter(function(h){{return h.pl!=null&&h.mv_eur!=null;}}).slice();
    sorted.sort(function(a,b){{
      var ra=(a.pl||0)/Math.max(Math.abs((a.mv_eur||0)-(a.pl||0)),1)*100;
      var rb=(b.pl||0)/Math.max(Math.abs((b.mv_eur||0)-(b.pl||0)),1)*100;
      return rb-ra;
    }});
    getVal=function(h){{return(h.pl||0)/Math.max(Math.abs((h.mv_eur||0)-(h.pl||0)),1)*100;}};
    fmtVal=function(v,isW){{return(v>=0?'+':'')+fmtEur.format(v)+' %';}};
  }}else if(gvMode==='day'){{
    /* Tagesrendite nicht direkt verf\u00fcgbar \u2013 zeige Gewichtung im Fonds */
    sorted=allH.filter(function(h){{return h.pct_nav!=null;}}).slice();
    sorted.sort(function(a,b){{return(b.pct_nav||0)-(a.pct_nav||0);}});
    getVal=function(h){{return h.pct_nav||0;}};
    fmtVal=function(v,isW){{return fmtEur.format(v)+' % NAV';}};
  }}else{{
    /* P&L gesamt in EUR */
    sorted=allH.filter(function(h){{return h.pl!=null;}}).slice();
    sorted.sort(function(a,b){{return(b.pl||0)-(a.pl||0);}});
    getVal=function(h){{return h.pl||0;}};
    fmtVal=function(v,isW){{return(v>=0?'+':'')+fmtInt.format(Math.round(v))+' \u20ac';}};
  }}
  var winners=sorted.slice(0,5);
  var losers=sorted.slice().reverse().slice(0,5);
  function makeRows(arr,isWinner){{
    if(!arr.length)return '<div style="color:#aaa;font-size:12px;padding:8px 0">Keine Daten</div>';
    return arr.map(function(h,i){{
      var nm=(h.name||'').substring(0,35),v=getVal(h);
      var plColor=isWinner?'#1a7c4a':'#c0392b';
      return'<div class="gv-row"><span class="gv-rank">'+(i+1)+'</span><span class="gv-name">'+esc(nm)+'</span><span class="gv-pl" style="color:'+plColor+'">'+fmtVal(v,isWinner)+'</span></div>';
    }}).join('');
  }}
  var dayNote=gvMode==='day'?'<div style="font-size:11px;color:#aaa;margin-bottom:10px">Tagesdaten nicht verf\u00fcgbar \u2014 zeigt gr\u00f6\u00dfte Positionen</div>':'';
  document.getElementById('gv-container').innerHTML=
    '<div><div class="gv-header winners">Gewinner</div>'+dayNote+makeRows(winners,true)+'</div>'+
    '<div><div class="gv-header losers">Verlierer</div>'+dayNote+makeRows(losers,false)+'</div>';
}}
document.getElementById('gv-tabs').addEventListener('click',function(e){{
  var btn=e.target.closest('.gv-tab');if(!btn)return;
  document.querySelectorAll('.gv-tab').forEach(function(b){{b.classList.remove('active');}});
  btn.classList.add('active');gvMode=btn.dataset.gv;renderGV();
}});
function renderHoldings(){{
  var f=FUNDS_DATA[currentFundIdx],holdings=(f.holdings||[]).slice();
  holdings.sort(function(a,b){{
    if(holdingsSortCol==='name'){{var va=(a.name||'').toLowerCase(),vb=(b.name||'').toLowerCase();return holdingsSortDir*(va<vb?-1:va>vb?1:0);}}
    if(holdingsSortCol==='sector'){{var va=(a.sector||'').toLowerCase(),vb=(b.sector||'').toLowerCase();return holdingsSortDir*(va<vb?-1:va>vb?1:0);}}
    var va=holdingsSortCol==='mv'?(a.mv_eur||0):holdingsSortCol==='pct'?(a.pct_nav||0):(a.pl||0);
    var vb=holdingsSortCol==='mv'?(b.mv_eur||0):holdingsSortCol==='pct'?(b.pct_nav||0):(b.pl||0);
    return holdingsSortDir*(vb-va);
  }});
  var rows='';
  holdings.forEach(function(h,i){{
    var plCls=h.pl==null?'':(h.pl>=0?' pos':' neg');
    rows+='<tr>'+
      '<td class="rank">'+(i+1)+'</td>'+
      '<td><div class="h-name">'+esc(h.name||'\u2014')+'</div><div class="h-isin">'+esc(h.isin||'')+'</div></td>'+
      '<td class="num">'+(h.mv_eur!=null?fmtInt.format(Math.round(h.mv_eur))+' \u20ac':'\u2014')+'</td>'+
      '<td class="num">'+(h.pct_nav!=null?fmtEur.format(h.pct_nav)+' %':'\u2014')+'</td>'+
      '<td class="num'+plCls+'">'+( h.pl!=null?(h.pl>=0?'+':'')+ fmtInt.format(Math.round(h.pl))+' \u20ac':'\u2014')+'</td>'+
      '<td>'+esc(h.sector||'\u2014')+'</td>'+
      '</tr>';
  }});
  document.getElementById('holdings-body').innerHTML=rows||'<tr><td colspan="6" style="text-align:center;color:#aaa;padding:24px">Keine Daten</td></tr>';
}}
document.getElementById('holdings-table').querySelector('thead').addEventListener('click',function(e){{
  var th=e.target.closest('th[data-col]');if(!th)return;
  var col=th.dataset.col;if(col==='rank')return;
  if(holdingsSortCol===col){{holdingsSortDir*=-1;}}else{{holdingsSortCol=col;holdingsSortDir=(col==='name'||col==='sector')?1:-1;}}
  renderHoldings();
}});
function groupBy(holdings,key,maxSlices){{
  var map={{}};
  holdings.forEach(function(h){{var k=h[key]||'Unbekannt';map[k]=(map[k]||0)+(h.mv_eur||0);}});
  var arr=Object.keys(map).map(function(k){{return {{label:k,value:map[k]}};}}); 
  arr.sort(function(a,b){{return b.value-a.value;}});
  if(arr.length>maxSlices){{var o=arr.splice(maxSlices),s=o.reduce(function(t,x){{return t+x.value;}},0);arr.push({{label:'Sonstige',value:s}});}}
  return arr;
}}
function buildDonut(canvasId,legendId,data,colors){{
  var ctx=document.getElementById(canvasId);if(!ctx)return null;
  var total=data.reduce(function(s,d){{return s+d.value;}},0);
  var chart=new Chart(ctx,{{
    type:'doughnut',
    data:{{
      labels:data.map(function(d){{return d.label;}}),
      datasets:[{{data:data.map(function(d){{return d.value;}}),backgroundColor:colors.slice(0,data.length),borderWidth:2,borderColor:'#fff'}}]
    }},
    options:{{
      responsive:true,maintainAspectRatio:false,
      plugins:{{
        legend:{{display:false}},
        tooltip:{{callbacks:{{label:function(ctx){{var pct=total>0?(ctx.parsed/total*100).toFixed(1):0;return ctx.label+': '+fmtInt.format(Math.round(ctx.parsed))+' \u20ac ('+pct+' %)';}}}}}}
      }}
    }}
  }});
  var lh='';
  data.forEach(function(d,i){{
    var pct=total>0?(d.value/total*100).toFixed(1):0;
    lh+='<div class="legend-item"><div class="legend-dot" style="background:'+(colors[i]||'#ccc')+'"></div><span class="legend-label">'+esc(d.label)+'</span><span class="legend-pct">'+pct+' %</span></div>';
  }});
  document.getElementById(legendId).innerHTML=lh;
  return chart;
}}
function renderAnalyse(){{
  if(chartSector){{chartSector.destroy();chartSector=null;}}
  if(chartCountry){{chartCountry.destroy();chartCountry=null;}}
  if(chartCurrency){{chartCurrency.destroy();chartCurrency=null;}}
  var h=FUNDS_DATA[currentFundIdx].holdings||[];
  chartSector=buildDonut('chart-sector','legend-sector',groupBy(h,'sector',8),DONUT_COLORS);
  chartCountry=buildDonut('chart-country','legend-country',groupBy(h,'country',8),DONUT_COLORS.slice(3).concat(DONUT_COLORS.slice(0,3)));
  chartCurrency=buildDonut('chart-currency','legend-currency',groupBy(h,'currency',6),DONUT_COLORS.slice(6).concat(DONUT_COLORS.slice(0,6)));
}}
function renderTransactions(){{
  var f=FUNDS_DATA[currentFundIdx];
  var changes=(CHANGES_HISTORY[f.id]||[]).slice().sort(function(a,b){{return(b.date||'').localeCompare(a.date||'');}});
  var typeMap={{'added':'Neukauf','removed':'Komplettverkauf','increased':'Aufstockung','decreased':'Teilverkauf'}};
  var typeColors={{'Neukauf':'#1a7c4a','Aufstockung':'#2563eb','Teilverkauf':'#d97706','Komplettverkauf':'#c0392b'}};
  var rows='';
  changes.forEach(function(c){{
    var label=typeMap[c.type]||c.type||'\u2014';
    var col=typeColors[label]||'#666';
    var mv=c.mv_eur!=null?fmtInt.format(Math.round(c.mv_eur))+' \u20ac':'\u2014';
    var pct=c.change_pct!=null?(c.change_pct>=0?'+':'')+fmtEur.format(c.change_pct)+' %':'\u2014';
    var pctCls=c.change_pct!=null?(c.change_pct>=0?' pos':' neg'):'';
    rows+='<tr>'+
      '<td class="date-cell">'+esc(c.date||'\u2014')+'</td>'+
      '<td><span class="badge" style="background:'+col+'20;color:'+col+'">'+esc(label)+'</span></td>'+
      '<td>'+esc(c.name||'\u2014')+'</td>'+
      '<td class="num'+pctCls+'">'+pct+'</td>'+
      '<td class="num">'+mv+'</td>'+
      '</tr>';
  }});
  document.getElementById('transactions-body').innerHTML=rows||'<tr><td colspan="5" style="text-align:center;color:#aaa;padding:24px">Keine Transaktionen</td></tr>';
}}
function renderRisk(){{
  var history=getHistory(),entries=history.slice(-252);
  function setRisk(v,d,s){{
    document.getElementById('risk-vol').textContent=v;
    document.getElementById('risk-drawdown').textContent=d;
    document.getElementById('risk-sharpe').textContent=s;
  }}
  if(entries.length<20){{setRisk('\u2014','\u2014','\u2014');return;}}
  var lr=[];
  for(var i=1;i<entries.length;i++){{if(entries[i-1].price&&entries[i].price)lr.push(Math.log(entries[i].price/entries[i-1].price));}}
  if(lr.length<5){{setRisk('\u2014','\u2014','\u2014');return;}}
  var mean=lr.reduce(function(s,r){{return s+r;}},0)/lr.length;
  var variance=lr.reduce(function(s,r){{return s+(r-mean)*(r-mean);}},0)/(lr.length-1);
  var vol=Math.sqrt(variance*252)*100;
  var peak=entries[0].price,maxDD=0;
  entries.forEach(function(e){{
    if(e.price>peak)peak=e.price;
    if(peak>0){{var dd=(peak-e.price)/peak;if(dd>maxDD)maxDD=dd;}}
  }});
  var annRet=mean*252*100,sharpe=vol>0?annRet/vol:null;
  setRisk(fmtEur.format(vol)+' %','-'+fmtEur.format(maxDD*100)+' %',sharpe!=null?fmtEur.format(sharpe):'\u2014');
}}
function renderMonthly(){{
  var history=getHistory();
  var now=new Date(),currentYear=now.getFullYear(),currentMonth=now.getMonth();
  var monthMap={{}};
  history.forEach(function(h){{
    var d=new Date(h.date+'T12:00:00');
    if(d.getFullYear()===currentYear){{
      var ym=h.date.slice(0,7);
      var ex=monthMap[ym];
      if(!ex){{monthMap[ym]=h;return;}}
      var hFull=(h.nav!=null&&h.perf_ytd!=null);
      var exFull=(ex.nav!=null&&ex.perf_ytd!=null);
      /* Prefer entries with complete data; among equal, take latest date */
      if(hFull&&!exFull){{monthMap[ym]=h;}}
      else if(hFull===exFull&&h.date>ex.date){{monthMap[ym]=h;}}
    }}
  }});
  /* Generate all months Jan \u2192 current month of current year */
  var allMonths=[];
  for(var m=0;m<=currentMonth;m++){{
    allMonths.push(currentYear+'-'+(m<9?'0':'')+(m+1));
  }}
  var mn=['Januar','Februar','M\u00e4rz','April','Mai','Juni','Juli','August','September','Oktober','November','Dezember'];
  var rows='';
  allMonths.forEach(function(ym,i){{
    var e=monthMap[ym]||null;
    var prevYm=i>0?allMonths[i-1]:null;
    var prev=prevYm?monthMap[prevYm]||null:null;
    var delta=(e&&prev&&prev.nav!=null&&e.nav!=null)?e.nav-prev.nav:null;
    var dcls=delta!=null?(delta>=0?' pos':' neg'):'';
    var dtxt=delta!=null?(delta>=0?'+':'')+ fmtInt.format(Math.round(Math.abs(delta)))+' \u20ac':'\u2014';
    var ytd=e?e.perf_ytd:null,ycls=ytd!=null?(ytd>=0?' pos':' neg'):'';
    var ytxt=ytd!=null?(ytd>=0?'+':'')+fmtEur.format(ytd)+' %':'\u2014';
    rows+='<tr>'+
      '<td>'+(mn[parseInt(ym.slice(5,7),10)-1]||ym)+'</td>'+
      '<td class="num">'+(e&&e.nav!=null?fmtInt.format(Math.round(e.nav))+' \u20ac':'\u2014')+'</td>'+
      '<td class="num">'+(e&&e.price!=null?fmtEur.format(e.price)+' \u20ac':'\u2014')+'</td>'+
      '<td class="num'+ycls+'">'+ytxt+'</td>'+
      '<td class="num'+dcls+'">'+dtxt+'</td>'+
      '</tr>';
  }});
  document.getElementById('monthly-body').innerHTML=rows||'<tr><td colspan="5" style="text-align:center;color:#aaa;padding:24px">Keine Daten</td></tr>';
}}
function buildNewsHtml(fundFilter){{
  var html='';
  Object.keys(NEWS_DATA).forEach(function(key){{
    var nd=NEWS_DATA[key];
    if(fundFilter&&nd.funds&&!nd.funds.includes(fundFilter))return;
    /* Sort articles newest-first */
    var articles=(nd.articles||[]).slice().sort(function(a,b){{
      return(b.pubDate||'').localeCompare(a.pubDate||'');
    }});
    if(!articles.length&&!nd.summary)return;
    var co=esc(nd.company||key);
    var sumRaw=nd.summary;
    var sumHeadline='',sumText='';
    if(typeof sumRaw==='string'){{sumText=sumRaw;}}
    else if(sumRaw&&typeof sumRaw==='object'){{sumHeadline=sumRaw.headline||'';sumText=sumRaw.text||'';}}
    var sumHtml='';
    if(sumHeadline)sumHtml+='<div class="news-headline">'+esc(sumHeadline)+'</div>';
    if(sumText)sumHtml+='<div class="news-summary">'+esc(sumText)+'</div>';
    var ah=articles.map(function(a){{
      return'<a class="article-link" href="'+esc(a.link||'#')+'" target="_blank" rel="noopener">'+esc(a.title||'')+'</a>';
    }}).join('');
    html+='<div class="news-card"><h3 class="news-company">'+co+'</h3>'+sumHtml+'<div class="articles-list">'+ah+'</div></div>';
  }});
  return html||'<p class="empty">Keine News verf\u00fcgbar.</p>';
}}
function renderNews(){{
  var f=FUNDS_DATA[currentFundIdx];
  document.getElementById('news-container').innerHTML=buildNewsHtml(f.id);
}}
function openNewsOverlay(){{
  document.getElementById('news-drawer-content').innerHTML=buildNewsHtml(null);
  document.getElementById('news-overlay').classList.add('open');
  document.body.style.overflow='hidden';
}}
function closeNewsOverlay(){{
  document.getElementById('news-overlay').classList.remove('open');
  document.body.style.overflow='';
}}
function overlayBgClick(e){{
  if(e.target===document.getElementById('news-overlay'))closeNewsOverlay();
}}
function renderRunLog(){{
  var rows='';
  RUN_LOG.slice(-5).reverse().forEach(function(e){{
    var sc=e.status==='ok'?'#1a7c4a':'#c0392b';
    rows+='<tr>'+
      '<td class="date-cell">'+esc(e.ts||'\u2014')+'</td>'+
      '<td><span class="badge" style="background:'+sc+'20;color:'+sc+'">'+esc(e.status||'\u2014')+'</span></td>'+
      '<td class="num">'+(e.funds!=null?e.funds:'\u2014')+'</td>'+
      '<td class="num">'+(e.holdings!=null?e.holdings:'\u2014')+'</td>'+
      '<td class="num">'+(e.news!=null?e.news:'\u2014')+'</td>'+
      '<td class="num">'+(e.summaries!=null?e.summaries:'\u2014')+'</td>'+
      '</tr>';
  }});
  document.getElementById('run-log-body').innerHTML=rows||'<tr><td colspan="6" style="text-align:center;color:#aaa;padding:16px">Keine Eintr\u00e4ge</td></tr>';
}}
document.querySelectorAll('.tab-btn').forEach(function(btn){{
  btn.addEventListener('click',function(){{switchTab(btn.dataset.tab);}});
}});
function maybeConfetti(){{
  var totalPl=FUNDS_DATA.reduce(function(s,f){{return s+(f.total_pl||0);}},0);
  if(totalPl>0&&typeof confetti!=='undefined'){{confetti({{particleCount:120,spread:70,origin:{{y:0.6}}}});}}
}}
renderSummaryBar();renderFundSelector();renderFundHeader();renderMetrics();renderOverview();updateTabColor();renderRunLog();
setTimeout(maybeConfetti,1200);
</script>
</body>
</html>'''
    return html


def git_push_file(token, repo, path, content_bytes, message, branch="main"):
    """Committed eine Datei direkt per GitHub API. Retry bei 5xx-Fehlern."""
    import base64
    import time as _time
    b64_content = base64.b64encode(content_bytes).decode()

    def _api(req, retries=3):
        for attempt in range(retries):
            try:
                with urlopen(req, timeout=30) as resp:
                    return json.loads(resp.read())
            except HTTPError as e:
                if e.code in (502, 503, 504) and attempt < retries - 1:
                    _time.sleep(3 * (attempt + 1))
                    continue
                raise
        raise RuntimeError("Unreachable")

    # Bestehende SHA holen (falls Datei existiert) — 5xx → Datei als neu behandeln
    sha = None
    try:
        sha_req = Request(
            f"https://api.github.com/repos/{repo}/contents/{path}?ref={branch}",
            headers={"Authorization": f"token {token}", "Accept": "application/vnd.github.v3+json"},
        )
        existing = _api(sha_req)
        sha = existing.get("sha")
    except HTTPError as e:
        if e.code != 404:
            print(f"  ⚠️  SHA-Abfrage {path}: HTTP {e.code} – fahre ohne SHA fort")
    except Exception as e:
        print(f"  ⚠️  SHA-Abfrage {path}: {e} – fahre ohne SHA fort")

    # Commit — bei 422 ohne SHA nochmal versuchen (SHA-Konflikt)
    for attempt in range(2):
        body = {"message": message, "content": b64_content, "branch": branch}
        if sha:
            body["sha"] = sha
        put_req = Request(
            f"https://api.github.com/repos/{repo}/contents/{path}",
            data=json.dumps(body).encode(),
            method="PUT",
            headers={
                "Authorization": f"token {token}",
                "Accept":        "application/vnd.github.v3+json",
                "Content-Type":  "application/json",
            },
        )
        try:
            result = _api(put_req)
            break
        except HTTPError as e:
            if e.code == 422 and sha and attempt == 0:
                # SHA veraltet → SHA weglassen (Datei neu anlegen)
                print(f"  ⚠️  SHA-Konflikt {path} – versuche ohne SHA")
                sha = None
                continue
            raise
    else:
        raise RuntimeError(f"Push fehlgeschlagen: {path}")

    commit_sha = result.get("commit", {}).get("sha", "")[:8]
    print(f"  ✅ Committed {path} → {commit_sha}")
    return result


def get_or_create_gh_pages(token, repo):
    """Aktiviert GitHub Pages (docs/ Ordner auf main)."""
    # Pages API
    req = Request(
        f"https://api.github.com/repos/{repo}/pages",
        data=json.dumps({"source": {"branch": "main", "path": "/docs"}}).encode(),
        method="POST",
        headers={
            "Authorization": f"token {token}",
            "Accept": "application/vnd.github.v3+json",
            "Content-Type": "application/json",
        },
    )
    try:
        with urlopen(req) as resp:
            pages = json.loads(resp.read())
            url = pages.get("html_url", "")
            print(f"  🌐 GitHub Pages aktiviert: {url}")
            return url
    except HTTPError as e:
        if e.code == 409:
            # Bereits aktiviert – URL lesen
            req2 = Request(
                f"https://api.github.com/repos/{repo}/pages",
                headers={"Authorization": f"token {token}", "Accept": "application/vnd.github.v3+json"},
            )
            with urlopen(req2) as resp:
                pages = json.loads(resp.read())
                url = pages.get("html_url", "")
                print(f"  🌐 GitHub Pages bereits aktiv: {url}")
                return url
        else:
            print(f"  ⚠️  Pages-Fehler {e.code}: {e.read()[:200]}")
            return ""


# ─── Prev Data laden/speichern ────────────────────────────────────────────────
def load_prev_data(token, repo, branch="main"):
    """Liest prev_data.json aus dem Repo (für Änderungserkennung)."""
    try:
        req = Request(
            f"https://api.github.com/repos/{repo}/contents/docs/prev_data.json?ref={branch}",
            headers={"Authorization": f"token {token}", "Accept": "application/vnd.github.v3+json"},
        )
        with urlopen(req) as resp:
            meta = json.loads(resp.read())
            content = base64.b64decode(meta["content"]).decode()
            return json.loads(content)
    except Exception as e:
        print(f"  ℹ️  Keine prev_data.json gefunden ({e}), starte frisch")
        return {}


def load_run_log(token, repo, branch="main"):
    """Liest den Run-Log aus docs/run_log.json (letzte N Runs)."""
    try:
        req = Request(
            f"https://api.github.com/repos/{repo}/contents/docs/run_log.json?ref={branch}",
            headers={"Authorization": f"token {token}", "Accept": "application/vnd.github.v3+json"},
        )
        with urlopen(req) as resp:
            meta = json.loads(resp.read())
            content = base64.b64decode(meta["content"]).decode()
            return json.loads(content)
    except Exception:
        return []


def load_nav_history(token, repo, branch="main"):
    """Liest akkumulierte NAV-Historie aus docs/nav_history.json."""
    try:
        req = Request(
            f"https://api.github.com/repos/{repo}/contents/docs/nav_history.json?ref={branch}",
            headers={"Authorization": f"token {token}", "Accept": "application/vnd.github.v3+json"},
        )
        with urlopen(req) as resp:
            meta = json.loads(resp.read())
            content = base64.b64decode(meta["content"]).decode()
            return json.loads(content)
    except Exception as e:
        print(f"  ℹ️  Keine nav_history.json ({e}), starte frisch")
        return {}


def load_json_from_github(token, repo, path, branch="main"):
    """Lädt eine beliebige JSON-Datei aus dem Repo."""
    try:
        req = Request(
            f"https://api.github.com/repos/{repo}/contents/{path}?ref={branch}",
            headers={"Authorization": f"token {token}", "Accept": "application/vnd.github.v3+json"},
        )
        with urlopen(req) as resp:
            meta = json.loads(resp.read())
            content = base64.b64decode(meta["content"]).decode()
            return json.loads(content)
    except Exception:
        return None


# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    github_token = os.environ.get("GITHUB_TOKEN", "")
    github_repo  = os.environ.get("GITHUB_REPOSITORY", "")
    RUN_MODE     = os.environ.get("RUN_MODE", "full").lower()  # "full" oder "news"

    print("=" * 60)
    print(f"🚀 IQAM Dashboard Update – {date.today()} [{RUN_MODE.upper()}]")
    print("=" * 60)

    # 2. Cached data laden
    prev_data = {}
    nav_history = {}
    run_log = []
    changes_history = {}
    holdings_prev = {}  # {fid: {"date": "...", "isins": [...]}}
    if github_token and github_repo:
        nav_history = load_nav_history(github_token, github_repo)
        run_log = load_run_log(github_token, github_repo)
        changes_history = load_json_from_github(github_token, github_repo, "docs/changes_history.json") or {}
        holdings_prev = load_json_from_github(github_token, github_repo, "docs/holdings_prev.json") or {}

    if RUN_MODE == "backfill":
        # ── Backfill: Alle historischen Holdings aus Outlook-Mails laden ───────
        print("\n🔄 BACKFILL-Modus: Lade alle historischen INVENTARLISTE-Mails…")
        access_token = get_access_token()
        holdings_history = backfill_holdings_history(access_token, {})

        # Changes-History aus vollständiger holdings_history neu aufbauen (inkl. Teilkäufe/-verkäufe)
        print("\n🔁 Baue Transaktionshistorie aus Holdings-History auf…")
        changes_history = {}
        for fid, snaps in holdings_history.items():
            changes_history[fid] = []
            existing_keys = set()
            sorted_dates = sorted(snaps.keys())
            for i in range(1, len(sorted_dates)):
                d_curr = sorted_dates[i]
                d_prev = sorted_dates[i - 1]
                curr_snap = {h["isin"]: h for h in snaps[d_curr] if h.get("isin")}
                prev_snap = {h["isin"]: h for h in snaps[d_prev] if h.get("isin")}
                for isin, h in curr_snap.items():
                    prev_h = prev_snap.get(isin)
                    if prev_h is None:
                        key = (isin, d_curr, "added")
                        if key not in existing_keys:
                            changes_history[fid].append({"date": d_curr, "type": "added", "isin": isin, "name": h.get("name",""), "mv_eur": h.get("mv_eur"), "qty": h.get("qty")})
                            existing_keys.add(key)
                    else:
                        prev_qty = prev_h.get("qty") or 0
                        curr_qty = h.get("qty") or 0
                        if prev_qty and curr_qty:
                            # Qty-basierte Erkennung (präzise)
                            if abs(curr_qty - prev_qty) / max(abs(prev_qty), 1) > 0.005:
                                diff_pct = (curr_qty - prev_qty) / abs(prev_qty) * 100
                                typ = "increased" if curr_qty > prev_qty else "decreased"
                                key = (isin, d_curr, typ)
                                if key not in existing_keys:
                                    price_per_share = round(h.get("mv_eur") / curr_qty, 4) if (h.get("mv_eur") and curr_qty) else None
                                    changes_history[fid].append({"date": d_curr, "type": typ, "isin": isin, "name": h.get("name",""), "mv_eur": h.get("mv_eur"), "qty": curr_qty, "prev_qty": prev_qty, "change_pct": round(diff_pct, 1), "price_per_share": price_per_share})
                                    existing_keys.add(key)
                        else:
                            # Fallback: mv_eur-Proxy wenn qty nicht im Excel (>10% Schwelle)
                            prev_mv = prev_h.get("mv_eur") or 0
                            curr_mv = h.get("mv_eur") or 0
                            if prev_mv and curr_mv and abs(curr_mv - prev_mv) / max(abs(prev_mv), 1) > 0.10:
                                diff_pct = (curr_mv - prev_mv) / abs(prev_mv) * 100
                                typ = "increased" if curr_mv > prev_mv else "decreased"
                                key = (isin, d_curr, typ)
                                if key not in existing_keys:
                                    changes_history[fid].append({"date": d_curr, "type": typ, "isin": isin, "name": h.get("name",""), "mv_eur": curr_mv, "change_pct": round(diff_pct, 1), "mv_proxy": True})
                                    existing_keys.add(key)
                for isin, h in prev_snap.items():
                    if isin not in curr_snap:
                        key = (isin, d_curr, "removed")
                        if key not in existing_keys:
                            changes_history[fid].append({"date": d_curr, "type": "removed", "isin": isin, "name": h.get("name",""), "mv_eur": h.get("mv_eur"), "qty": h.get("qty")})
                            existing_keys.add(key)
            total_ch = len(changes_history[fid])
            print(f"  📊 {fid}: {len(sorted_dates)} Tage, {total_ch} Transaktionen erkannt")

        # Pushen — changes_history zuerst (wichtig!), holdings_history nicht pushen (zu groß)
        # Stattdessen: letzten Snapshot als holdings_prev speichern (für zukünftige tägliche Vergleiche)
        today_str = date.today().isoformat()
        holdings_prev_new = {}
        for fid, snaps in holdings_history.items():
            if snaps:
                last_date = sorted(snaps.keys())[-1]
                holdings_prev_new[fid] = {
                    "date": last_date,
                    "isins": {h["isin"]: {"name": h.get("name",""), "mv_eur": h.get("mv_eur")}
                              for h in snaps[last_date] if h.get("isin")}
                }
        # NAV-History aus INVENTARBLATT-Mails backfillen
        print("\n📈 Backfille NAV-History aus INVENTARBLATT-Mails…")
        nav_history = backfill_nav_history_from_emails(access_token, nav_history)

        if github_token and github_repo:
            print("\n📤 Pushe changes_history.json, holdings_prev.json und nav_history.json…")
            git_push_file(github_token, github_repo, "docs/changes_history.json",
                         json.dumps(changes_history, ensure_ascii=False).encode("utf-8"),
                         f"Backfill changes history {today_str}")
            git_push_file(github_token, github_repo, "docs/holdings_prev.json",
                         json.dumps(holdings_prev_new, ensure_ascii=False).encode("utf-8"),
                         f"Backfill holdings prev {today_str}")
            git_push_file(github_token, github_repo, "docs/nav_history.json",
                         json.dumps(nav_history, ensure_ascii=False).encode("utf-8"),
                         f"Backfill nav history {today_str}")
        print("\n✅ Backfill abgeschlossen!")
        return

    if RUN_MODE == "news":
        # ── News-only Run: Fondsdaten aus Cache laden ──────────────────────────
        print("\n📂 News-only Run: Lade gecachte Fondsdaten…")
        cached = load_json_from_github(github_token, github_repo, "docs/prev_data.json")
        if not cached:
            print("❌ Keine gecachten Fondsdaten gefunden. Bitte zuerst Full-Run ausführen.")
            sys.exit(1)
        funds_data = list(cached.values())
        print(f"  ✅ {len(funds_data)} Fonds aus Cache geladen")
    else:
        # ── Full Run: Fondsdaten frisch aus Email holen ────────────────────────
        # 1. MS Token
        access_token = get_access_token()

        # 2. Prev data für Änderungserkennung
        if github_token and github_repo:
            prev_data = load_prev_data(github_token, github_repo)

        # 3. Mails finden
        fund_mails = find_latest_emails(access_token)

    # 4. Pro Fund Excel laden + parsen (nur Full-Run)
    if RUN_MODE != "news":
        funds_data = []
        for fund_meta in FUNDS:
            fid        = fund_meta["id"]
            mail_entry = fund_mails.get(fid, {})
            mail_blatt = mail_entry.get("blatt")
            mail_liste = mail_entry.get("liste")
    
            if not mail_blatt and not mail_liste:
                print(f"⚠️  Keine Mail für Fund {fid} gefunden!")
                prev_fund = prev_data.get(fid, {})
                if prev_fund:
                    print(f"   → Verwende gestrige Daten für {fid}")
                    funds_data.append({**fund_meta, **prev_fund, "changes": {}})
                continue
    
            fund_parsed = {}
    
            # INVENTARBLATT → NAV, Preis, YTD, FY
            if mail_blatt:
                print(f"\n📋 Fund {fid} BLATT: {mail_blatt['subject'][:55]}")
                xlsx_bytes, filename = download_attachment(access_token, mail_blatt["id"], ".xlsx")
                if xlsx_bytes:
                    print(f"   Parsing {filename}…")
                    try:
                        blatt_data = parse_excel(xlsx_bytes, fid)
                        fund_parsed.update(blatt_data)
                    except Exception as e:
                        print(f"   ❌ BLATT Parse-Fehler: {e}")
                        traceback.print_exc()
    
            # INVENTARLISTE → Holdings, Länder, Währungen, Sektoren
            mail_for_holdings = mail_liste or mail_blatt  # Fallback auf BLATT wenn keine LISTE
            if mail_for_holdings:
                print(f"\n📊 Fund {fid} LISTE: {mail_for_holdings['subject'][:55]}")
                xlsx_bytes, filename = download_attachment(access_token, mail_for_holdings["id"], ".xlsx")
                if xlsx_bytes:
                    print(f"   Parsing {filename}…")
                    try:
                        liste_data = parse_excel(xlsx_bytes, fid)
                        # Merge: LISTE-Daten überschreiben nur Holdings/Allokation, nicht NAV
                        for key in ["holdings", "countries", "currencies", "sectors"]:
                            if key in liste_data:
                                fund_parsed[key] = liste_data[key]
                        # NAV aus BLATT bevorzugen — nur übernehmen wenn noch nicht gesetzt
                        for key in ["nav", "nav_per_share", "shares", "perf_ytd", "perf_fy", "report_date"]:
                            if key not in fund_parsed and key in liste_data:
                                fund_parsed[key] = liste_data[key]
                    except Exception as e:
                        print(f"   ❌ LISTE Parse-Fehler: {e}")
                        traceback.print_exc()
    
            # Vortags-Preis: aus prev_data wenn run_date vor heute gesetzt, sonst nav_history
            # run_date wird von uns gesetzt (nicht aus Excel) → zuverlässig unabhängig von NAV-Datum
            prev_fund = prev_data.get(fid, {})
            today_iso = date.today().isoformat()
            prev_run_date = prev_fund.get("run_date", "")
            if prev_fund and prev_run_date and prev_run_date < today_iso:
                # Vorheriger Run war gestern (oder früher) → direkt als Baseline
                nav_ps_prev = prev_fund.get("nav_per_share")
            else:
                # Kein valides prev_data → nav_history als Fallback (nur real gemessene Punkte)
                hist_entries = [h for h in nav_history.get(fid, []) if h["date"] < today_iso and h.get("source") == "measured"]
                if not hist_entries:
                    # Alle Punkte (auch Seed-Punkte) falls keine gemessenen vorhanden
                    hist_entries = [h for h in nav_history.get(fid, []) if h["date"] < today_iso]
                nav_ps_prev = hist_entries[-1]["price"] if hist_entries else None
            fund_parsed["nav_per_share_prev"] = nav_ps_prev
            fund_parsed["nav_prev"] = prev_fund.get("nav") if prev_fund else None
    
            # Price history aufbauen
            fund_parsed["price_history"] = build_price_history(
                fund_parsed.get("nav_per_share"),
                fund_parsed.get("perf_ytd"),
                fund_parsed.get("perf_fy"),
                nav_ps_prev,
            )
    
            # Änderungen erkennen
            prev_holdings = prev_fund.get("holdings", []) if prev_fund else []
            changes = detect_changes(fund_parsed.get("holdings", []), prev_holdings)
            changes["date_prev"] = prev_fund.get("report_date") if prev_fund else None
            fund_parsed["changes"] = changes

            # Tägliche Änderungen vs. letztem Snapshot erkennen und changes_history aktualisieren
            today_str = date.today().isoformat()
            curr_map = {h.get("isin"): h for h in fund_parsed.get("holdings", [])
                        if h.get("isin") and h["isin"] not in ("None","")}
            prev_snap_entry = holdings_prev.get(fid, {})
            prev_isin_map = prev_snap_entry.get("isins", {})
            prev_date = prev_snap_entry.get("date", "")

            if fid not in changes_history:
                changes_history[fid] = []
            existing_keys = {(e["isin"], e["date"], e["type"]) for e in changes_history[fid]}

            if prev_isin_map and prev_date and prev_date < today_str:
                for isin, h in curr_map.items():
                    prev_info = prev_isin_map.get(isin)
                    if prev_info is None:
                        # Neukauf (komplett neu)
                        key = (isin, today_str, "added")
                        if key not in existing_keys:
                            changes_history[fid].append({"date": today_str, "type": "added", "isin": isin, "name": h.get("name",""), "mv_eur": h.get("mv_eur"), "qty": h.get("qty")})
                            existing_keys.add(key)
                    else:
                        # Teilkauf / Teilverkauf (Position bleibt, Menge ändert sich)
                        prev_qty = prev_info.get("qty") or 0
                        curr_qty = h.get("qty") or 0
                        if prev_qty and curr_qty:
                            # Qty-basierte Erkennung (präzise)
                            if abs(curr_qty - prev_qty) / max(abs(prev_qty), 1) > 0.005:
                                diff_pct = (curr_qty - prev_qty) / abs(prev_qty) * 100
                                typ = "increased" if curr_qty > prev_qty else "decreased"
                                key = (isin, today_str, typ)
                                if key not in existing_keys:
                                    price_per_share = round(h.get("mv_eur") / curr_qty, 4) if (h.get("mv_eur") and curr_qty) else None
                                    changes_history[fid].append({"date": today_str, "type": typ, "isin": isin, "name": h.get("name",""), "mv_eur": h.get("mv_eur"), "qty": curr_qty, "prev_qty": prev_qty, "change_pct": round(diff_pct, 1), "price_per_share": price_per_share})
                                    existing_keys.add(key)
                        else:
                            # Fallback: mv_eur-Proxy wenn qty nicht im Excel vorhanden (>10% Schwelle)
                            prev_mv = prev_info.get("mv_eur") or 0
                            curr_mv = h.get("mv_eur") or 0
                            if prev_mv and curr_mv and abs(curr_mv - prev_mv) / max(abs(prev_mv), 1) > 0.10:
                                diff_pct = (curr_mv - prev_mv) / abs(prev_mv) * 100
                                typ = "increased" if curr_mv > prev_mv else "decreased"
                                key = (isin, today_str, typ)
                                if key not in existing_keys:
                                    changes_history[fid].append({"date": today_str, "type": typ, "isin": isin, "name": h.get("name",""), "mv_eur": curr_mv, "change_pct": round(diff_pct, 1), "mv_proxy": True})
                                    existing_keys.add(key)
                for isin, info in prev_isin_map.items():
                    if isin not in curr_map:
                        # Komplettverkauf
                        key = (isin, today_str, "removed")
                        if key not in existing_keys:
                            changes_history[fid].append({"date": today_str, "type": "removed", "isin": isin, "name": info.get("name",""), "mv_eur": info.get("mv_eur"), "qty": info.get("qty")})
                            existing_keys.add(key)

            # Aktuellen Snapshot als neues holdings_prev speichern (inkl. qty für Teilkauf/-verkauf-Erkennung)
            holdings_prev[fid] = {
                "date": today_str,
                "isins": {isin: {"name": h.get("name",""), "mv_eur": h.get("mv_eur"), "qty": h.get("qty")} for isin, h in curr_map.items()}
            }
    
            # Vortags-Holdings für Tagesvergleich (lean – nur nötige Felder)
            fund_parsed["prev_holdings"] = [
                {"isin": h.get("isin"), "name": h.get("name"), "mv_eur": h.get("mv_eur"),
                 "pl": h.get("pl"), "qty": h.get("qty")}
                for h in prev_holdings if h.get("isin") and h["isin"] not in ("None", "")
            ]
    
            # KPIs berechnen
            fund_parsed = compute_kpis(fund_parsed)
    
            # Merge mit Fund-Metadaten
            funds_data.append({**fund_meta, **fund_parsed})
    
            print(f"   ✅ {fid}: NAV {fund_parsed.get('nav',0)/1e6:.2f} Mio., "
                  f"Preis {fund_parsed.get('nav_per_share',0):.4f}, "
                  f"YTD {fund_parsed.get('perf_ytd',0):.2f}%")

    if not funds_data:
        print("❌ Keine Daten gefunden. Abbruch.")
        sys.exit(1)

    # 4b. NAV-Historie aktualisieren (nur Full-Run)
    today_str = date.today().isoformat()
    if RUN_MODE != "news":
        for fund in funds_data:
            fid      = fund["id"]
            price    = fund.get("nav_per_share")
            nav      = fund.get("nav")
            shares   = fund.get("shares")
            perf_ytd = fund.get("perf_ytd")
            # Fallback: Nettovermögen = Preis × Anteile wenn direkt nicht gefunden
            if nav is None and price and shares:
                nav = float(price) * float(shares)
            if price and price > 0:
                if fid not in nav_history:
                    nav_history[fid] = []
                # BVI-Seed-Punkte hinzufügen (GJ-Start, YTD-Start)
                for ph_point in fund.get("price_history", []):
                    if not any(h["date"] == ph_point["date"] for h in nav_history[fid]):
                        nav_history[fid].append({"date": ph_point["date"], "price": ph_point["price"]})
                # Heutigen Datenpunkt hinzufügen (als "measured" markiert, für zuverlässige Baseline)
                if not any(h["date"] == today_str for h in nav_history[fid]):
                    nav_history[fid].append({
                        "date": today_str,
                        "price": round(price, 4),
                        "nav": round(nav, 2) if nav else None,
                        "perf_ytd": round(float(perf_ytd), 4) if perf_ytd is not None else None,
                        "source": "measured",
                    })
                nav_history[fid].sort(key=lambda x: x["date"])
                print(f"  📈 {fid} NAV-Historie: {len(nav_history[fid])} Punkte")

    # 5. News fetchen (max 2× pro Tag)
    today_str = date.today().isoformat()
    news_runs_today = sum(
        1 for e in run_log
        if e.get("ts", "").startswith(today_str) and e.get("news", 0) > 0
    )
    prev_news_data = {}
    if github_token and github_repo:
        prev_news_data = load_json_from_github(github_token, github_repo, "docs/news_data.json") or {}

    if news_runs_today >= 2 and prev_news_data:
        print(f"\n⏭️  News bereits {news_runs_today}× heute aktualisiert – überspringe News-Fetch.")
        news_data = prev_news_data
    else:
        companies_for_news = {}
        for fund in funds_data:
            fid = fund["id"]
            for h in fund.get("holdings", []):
                isin = (h.get("isin") or "").strip()
                name = (h.get("name") or "").strip()
                if not name or name in ("None", ""):
                    continue
                key = isin if (isin and isin not in ("None", "")) else name
                if key not in companies_for_news:
                    companies_for_news[key] = {"name": name, "funds": [], "mv": 0}
                companies_for_news[key]["mv"] += h.get("mv_eur") or 0
                if fid not in companies_for_news[key]["funds"]:
                    companies_for_news[key]["funds"].append(fid)
        companies_for_news = dict(sorted(companies_for_news.items(), key=lambda x: x[1]["mv"], reverse=True))
        anthropic_key = os.environ.get("ANTHROPIC_API_KEY", "")
        news_data = fetch_all_news(companies_for_news, anthropic_key=anthropic_key, prev_news_data=prev_news_data, max_summaries=10, max_wall_seconds=240)
        # Fallback: wenn Fetch nichts liefert (z.B. Google blockiert GitHub IPs) → Altdaten behalten
        if not news_data and prev_news_data:
            print("  ↩️  News-Fetch lieferte nichts – verwende gecachte Altdaten.")
            news_data = prev_news_data

    # 5b. Run-Log Eintrag erstellen
    run_entry = {
        "ts": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "status": "success",
        "funds": len(funds_data),
        "holdings": sum(len(f.get("holdings", [])) for f in funds_data),
        "news": len(news_data),
        "summaries": sum(1 for v in news_data.values() if v.get("summary")),
        "aum": sum(f.get("nav", 0) for f in funds_data),
    }
    run_log.insert(0, run_entry)
    run_log = run_log[:30]  # max 30 Einträge

    # 6. Dashboard generieren
    updated_at = datetime.now().strftime("%d.%m.%Y %H:%M UTC")
    print(f"\n🔨 Generiere Dashboard ({updated_at})…")
    html = generate_html(funds_data, updated_at, nav_history=nav_history, news_data=news_data, run_log=run_log, changes_history=changes_history)
    data_json = json.dumps(
        [{k: v for k, v in f.items() if k != "holdings"} | {"holdings": f.get("holdings", [])}
         for f in funds_data],
        ensure_ascii=False, indent=2
    )

    # 7. In GitHub pushen
    if github_token and github_repo:
        print(f"\n📤 Push zu {github_repo}…")
        today_str = date.today().isoformat()
        git_push_file(github_token, github_repo, "docs/index.html",
                     html.encode("utf-8"),
                     f"Dashboard update {today_str}")
        # Jekyll-Verarbeitung deaktivieren (verhindert Build-Fehler durch {{ }} in JSON-Daten)
        git_push_file(github_token, github_repo, "docs/.nojekyll",
                     b"",
                     "Disable Jekyll")
        git_push_file(github_token, github_repo, "docs/dashboard_data.json",
                     data_json.encode("utf-8"),
                     f"Data update {today_str}")
        # News-Daten speichern — nur wenn vorhanden (verhindert Überschreiben mit leerem Dict)
        if news_data:
            git_push_file(github_token, github_repo, "docs/news_data.json",
                         json.dumps(news_data, ensure_ascii=False).encode("utf-8"),
                         f"News update {today_str}")
        # Nur beim Full-Run: Prev-Data, NAV-Historie speichern
        if RUN_MODE != "news":
            prev_save = {f["id"]: {**{k: v for k, v in f.items() if k not in ("changes",)}, "run_date": today_str}
                         for f in funds_data}
            git_push_file(github_token, github_repo, "docs/prev_data.json",
                         json.dumps(prev_save, ensure_ascii=False).encode("utf-8"),
                         f"Prev data {today_str}")
            git_push_file(github_token, github_repo, "docs/nav_history.json",
                         json.dumps(nav_history, ensure_ascii=False).encode("utf-8"),
                         f"NAV history {today_str}")
            git_push_file(github_token, github_repo, "docs/changes_history.json",
                         json.dumps(changes_history, ensure_ascii=False).encode("utf-8"),
                         f"Changes history {today_str}")
            git_push_file(github_token, github_repo, "docs/holdings_prev.json",
                         json.dumps(holdings_prev, ensure_ascii=False).encode("utf-8"),
                         f"Holdings prev {today_str}")
        git_push_file(github_token, github_repo, "docs/run_log.json",
                     json.dumps(run_log, ensure_ascii=False).encode("utf-8"),
                     f"Run log {today_str}")
        # GitHub Pages aktivieren (idempotent)
        pages_url = get_or_create_gh_pages(github_token, github_repo)
        if pages_url:
            print(f"\n🌐 Dashboard URL: {pages_url}")
    else:
        # Lokaler Modus: Dateien schreiben
        print("\n💾 Lokaler Modus (kein GITHUB_TOKEN gesetzt)")
        with open("docs/index.html", "w", encoding="utf-8") as f:
            f.write(html)
        with open("docs/dashboard_data.json", "w", encoding="utf-8") as f:
            f.write(data_json)
        print("   → docs/index.html")
        print("   → docs/dashboard_data.json")

    print("\n✅ Dashboard-Update abgeschlossen!")


if __name__ == "__main__":
    main()
